# lumber_track/views.py
import json
import re
from datetime import datetime
from django.db import models
from django.db.models import Q, Sum
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import connection, IntegrityError

from .models import (
    ProductType, WoodSpecies, QualityGrade, ProductName,
    UnitDimension, LumberDimension, ProductItem, Document, DocumentItem, StorageLocation
)


# ========== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ==========
def get_item_key(item):
    """Формирует уникальный ключ для позиции"""
    return f"{item.product_name_id}_{item.species_id}_{item.grade_id}_{item.lumber_dim_id}_{item.unit_dim_id}"


def create_item_dict(item):
    """Создает словарь с данными позиции"""
    if item.lumber_dim:
        dimension_display = f"{item.lumber_dim.thickness}×{item.lumber_dim.width}×{item.lumber_dim.length}"
        volume = item.lumber_dim.volume_m3
        area = item.lumber_dim.area_m2
    elif item.unit_dim:
        dimension_display = f"{item.unit_dim.length}×{item.unit_dim.width}×{item.unit_dim.height}"
        volume = 0
        area = 0
    else:
        dimension_display = "—"
        volume = 0
        area = 0

    return {
        'product_name': item.product_name.name,
        'species': item.species.name,
        'grade': item.grade.code,
        'dimension_display': dimension_display,
        'quantity': 0,
        'volume': 0,
        'area': 0,
    }


def get_available_stocks_with_details(location_id=None, exclude_document_id=None):
    """Возвращает позиции с остатками и деталями для селекта расхода"""
    query = """
        WITH stock_calc AS (
            SELECT 
                ROW_NUMBER() OVER (ORDER BY pn.name, ws.name, qg.code) as id,
                di.product_name_id,
                pn.name as product_name,
                di.species_id,
                ws.name as species,
                di.grade_id,
                qg.code as grade,
                di.lumber_dim_id,
                ld.thickness,
                ld.width,
                ld.length,
                di.unit_dim_id,
                ud.length as unit_length,
                ud.width as unit_width,
                ud.height as unit_height,
                COALESCE(SUM(CASE WHEN d.doc_type = 1 THEN di.quantity ELSE 0 END), 0) +
                COALESCE(SUM(CASE WHEN d.doc_type = 2 THEN di.quantity ELSE 0 END), 0) -
                COALESCE(SUM(CASE WHEN d.doc_type = 3 THEN di.quantity ELSE 0 END), 0) as balance
            FROM lumber_track_documentitem di
            JOIN lumber_track_document d ON di.document_id = d.id
            JOIN lumber_track_productname pn ON di.product_name_id = pn.id
            JOIN lumber_track_woodspecies ws ON di.species_id = ws.id
            JOIN lumber_track_qualitygrade qg ON di.grade_id = qg.id
            LEFT JOIN lumber_track_lumberdimension ld ON di.lumber_dim_id = ld.id
            LEFT JOIN lumber_track_unitdimension ud ON di.unit_dim_id = ud.id
            WHERE d.doc_date <= date('now')
    """
    if location_id:
        query += f" AND d.location_id = {location_id}"
    if exclude_document_id:
        query += f" AND d.id != {exclude_document_id}"
    query += """
            GROUP BY 
                di.product_name_id, pn.name,
                di.species_id, ws.name,
                di.grade_id, qg.code,
                di.lumber_dim_id, ld.thickness, ld.width, ld.length,
                di.unit_dim_id, ud.length, ud.width, ud.height
            HAVING balance > 0
        )
        SELECT 
            id,
            product_name_id,
            product_name,
            species_id,
            species,
            grade_id,
            grade,
            lumber_dim_id,
            thickness,
            width,
            length,
            unit_dim_id,
            unit_length,
            unit_width,
            unit_height,
            balance,
            CASE 
                WHEN lumber_dim_id IS NOT NULL THEN 
                    thickness || '-' || width || '-' || length || ' мм'
                ELSE 
                    unit_length || '-' || unit_width || '-' || unit_height || ' мм'
            END as dimension_display,
            CASE 
                WHEN lumber_dim_id IS NOT NULL THEN 'lumber'
                ELSE 'unit'
            END as dimension_type,
            COALESCE(lumber_dim_id, unit_dim_id) as dimension_id
        FROM stock_calc
        ORDER BY product_name, species, grade, dimension_display
    """
    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        results = [dict(zip(columns, row)) for row in cursor.fetchall()]
    return results


def export_to_excel(data, title, date_from, date_to):
    """Экспорт отчета в Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel ограничение на длину имени листа

    headers = ['Наименование', 'Порода', 'Категория', 'Размер', 'Количество (шт)', 'Объем (м³)', 'Площадь (м²)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

    for row, item in enumerate(data, 2):
        ws.cell(row=row, column=1, value=item['product_name'])
        ws.cell(row=row, column=2, value=item['species'])
        ws.cell(row=row, column=3, value=item['grade'])
        ws.cell(row=row, column=4, value=item['dimension_display'])
        ws.cell(row=row, column=5, value=float(item['total_quantity']))
        ws.cell(row=row, column=6, value=float(item['total_volume']))
        ws.cell(row=row, column=7, value=float(item['total_area']))

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{title}_{date_from}_{date_to}.xlsx"'
    wb.save(response)
    return response


# ========== ГЛАВНАЯ И СПРАВОЧНИКИ ==========
def home_view(request):
    return render(request, 'lumber_track/home.html')


def directories_view(request):
    return render(request, 'lumber_track/directories.html')


# ========== ТИПЫ ПРОДУКЦИИ ==========
def producttype_list(request):
    items = ProductType.objects.all().order_by('name')
    context = {
        'title': 'Типы продукции',
        'items': items,
        'create_url': '/directories/producttype/create/',
        'delete_url': 'lumber_track:producttype_delete',
    }
    return render(request, 'lumber_track/directory_table.html', context)


def producttype_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            ProductType.objects.create(name=name)
    return redirect('lumber_track:producttype_list')


def producttype_edit(request, pk):
    item = get_object_or_404(ProductType, pk=pk)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            item.name = name
            item.save()
    return redirect('lumber_track:producttype_list')


def producttype_delete(request, pk):
    item = get_object_or_404(ProductType, pk=pk)
    item.delete()
    return redirect('lumber_track:producttype_list')


def producttype_data(request, pk):
    item = get_object_or_404(ProductType, pk=pk)
    return JsonResponse({'name': item.name})


# ========== ПОРОДЫ ДРЕВЕСИНЫ ==========
def woodspecies_list(request):
    items = WoodSpecies.objects.all().order_by('name')
    context = {
        'title': 'Породы древесины',
        'items': items,
        'create_url': '/directories/woodspecies/create/',
        'delete_url': 'lumber_track:woodspecies_delete',
    }
    return render(request, 'lumber_track/directory_table.html', context)


def woodspecies_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            WoodSpecies.objects.create(name=name)
    return redirect('lumber_track:woodspecies_list')


def woodspecies_edit(request, pk):
    item = get_object_or_404(WoodSpecies, pk=pk)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            item.name = name
            item.save()
    return redirect('lumber_track:woodspecies_list')


def woodspecies_delete(request, pk):
    item = get_object_or_404(WoodSpecies, pk=pk)
    item.delete()
    return redirect('lumber_track:woodspecies_list')


def woodspecies_data(request, pk):
    item = get_object_or_404(WoodSpecies, pk=pk)
    return JsonResponse({'name': item.name})


# ========== КАТЕГОРИИ КАЧЕСТВА ==========
def qualitygrade_list(request):
    items = QualityGrade.objects.all().order_by('code')
    context = {
        'title': 'Категории качества',
        'items': items,
        'create_url': '/directories/qualitygrade/create/',
        'delete_url': 'lumber_track:qualitygrade_delete',
    }
    return render(request, 'lumber_track/directory_table.html', context)


def qualitygrade_create(request):
    if request.method == 'POST':
        code = request.POST.get('name', '').strip().upper()
        if code:
            QualityGrade.objects.create(code=code)
    return redirect('lumber_track:qualitygrade_list')


def qualitygrade_edit(request, pk):
    item = get_object_or_404(QualityGrade, pk=pk)
    if request.method == 'POST':
        code = request.POST.get('name', '').strip().upper()
        if code:
            item.code = code
            item.save()
    return redirect('lumber_track:qualitygrade_list')


def qualitygrade_delete(request, pk):
    item = get_object_or_404(QualityGrade, pk=pk)
    item.delete()
    return redirect('lumber_track:qualitygrade_list')


def qualitygrade_data(request, pk):
    item = get_object_or_404(QualityGrade, pk=pk)
    return JsonResponse({'name': item.code})


# ========== НАИМЕНОВАНИЯ ИЗДЕЛИЙ ==========
def productname_list(request):
    items = ProductName.objects.all().select_related('product_type').order_by('name')
    product_types = ProductType.objects.all().order_by('name')
    context = {
        'title': 'Наименования изделий',
        'items': items,
        'product_types': product_types,
        'create_url': '/directories/productname/create/',
        'delete_url': 'lumber_track:productname_delete',
    }
    return render(request, 'lumber_track/directory_productname.html', context)


def productname_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        product_type_id = request.POST.get('product_type')
        if name and product_type_id:
            product_type = get_object_or_404(ProductType, pk=product_type_id)
            ProductName.objects.create(name=name, product_type=product_type)
    return redirect('lumber_track:productname_list')


def productname_edit(request, pk):
    item = get_object_or_404(ProductName, pk=pk)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        product_type_id = request.POST.get('product_type')
        if name and product_type_id:
            item.name = name
            item.product_type_id = product_type_id
            item.save()
    return redirect('lumber_track:productname_list')


def productname_delete(request, pk):
    item = get_object_or_404(ProductName, pk=pk)
    item.delete()
    return redirect('lumber_track:productname_list')


def productname_data(request, pk):
    item = get_object_or_404(ProductName, pk=pk)
    return JsonResponse({
        'name': item.name,
        'product_type_id': item.product_type_id
    })


# ========== РАЗМЕРЫ ШТУЧНЫХ ИЗДЕЛИЙ ==========
def unitdimension_list(request):
    items = UnitDimension.objects.all().order_by('length', 'width', 'height')
    context = {
        'title': 'Размеры изделий',
        'items': items,
        'create_url': '/directories/unitdimension/create/',
        'delete_url': 'lumber_track:unitdimension_delete',
    }
    return render(request, 'lumber_track/directory_unitdimension.html', context)


def unitdimension_create(request):
    if request.method == 'POST':
        length = request.POST.get('length', '').strip()
        width = request.POST.get('width', '').strip()
        height = request.POST.get('height', '').strip()
        if length and width and height:
            try:
                UnitDimension.objects.create(
                    length=int(length),
                    width=int(width),
                    height=int(height)
                )
            except (ValueError, IntegrityError):
                pass
    return redirect('lumber_track:unitdimension_list')


def unitdimension_edit(request, pk):
    item = get_object_or_404(UnitDimension, pk=pk)
    if request.method == 'POST':
        length = request.POST.get('length', '').strip()
        width = request.POST.get('width', '').strip()
        height = request.POST.get('height', '').strip()
        if length and width and height:
            try:
                item.length = int(length)
                item.width = int(width)
                item.height = int(height)
                item.save()
            except ValueError:
                pass
    return redirect('lumber_track:unitdimension_list')


def unitdimension_delete(request, pk):
    item = get_object_or_404(UnitDimension, pk=pk)
    item.delete()
    return redirect('lumber_track:unitdimension_list')


def unitdimension_data(request, pk):
    item = get_object_or_404(UnitDimension, pk=pk)
    return JsonResponse({
        'length': item.length,
        'width': item.width,
        'height': item.height,
        'display': str(item)
    })


# ========== РАЗМЕРЫ ПОГОНАЖА ==========
def lumberdimension_list(request):
    items = LumberDimension.objects.all().order_by('thickness', 'width', 'length')
    context = {
        'title': 'Размеры погонажа',
        'items': items,
        'create_url': '/directories/lumberdimension/create/',
        'delete_url': 'lumber_track:lumberdimension_delete',
    }
    return render(request, 'lumber_track/directory_lumberdimension.html', context)


def lumberdimension_create(request):
    if request.method == 'POST':
        thickness = request.POST.get('thickness', '').strip()
        width = request.POST.get('width', '').strip()
        length = request.POST.get('length', '').strip()
        if thickness and width and length:
            try:
                LumberDimension.objects.create(
                    thickness=int(thickness),
                    width=int(width),
                    length=int(length)
                )
            except (ValueError, IntegrityError):
                pass
    return redirect('lumber_track:lumberdimension_list')


def lumberdimension_edit(request, pk):
    item = get_object_or_404(LumberDimension, pk=pk)
    if request.method == 'POST':
        thickness = request.POST.get('thickness', '').strip()
        width = request.POST.get('width', '').strip()
        length = request.POST.get('length', '').strip()
        if thickness and width and length:
            try:
                item.thickness = int(thickness)
                item.width = int(width)
                item.length = int(length)
                item.save()
            except ValueError:
                pass
    return redirect('lumber_track:lumberdimension_list')


def lumberdimension_delete(request, pk):
    item = get_object_or_404(LumberDimension, pk=pk)
    item.delete()
    return redirect('lumber_track:lumberdimension_list')


def lumberdimension_data(request, pk):
    item = get_object_or_404(LumberDimension, pk=pk)
    return JsonResponse({
        'thickness': item.thickness,
        'width': item.width,
        'length': item.length,
        'display': str(item),
        'volume': round(item.volume_m3, 6),
        'area': round(item.area_m2, 3)
    })


# ========== СПРАВОЧНИК ИЗДЕЛИЙ ==========
def productitem_list(request):
    items = ProductItem.objects.all().select_related(
        'product_name', 'species', 'grade', 'lumber_dim', 'unit_dim'
    ).order_by('-created_at')
    context = {
        'title': 'Справочник изделий',
        'items': items,
        'product_names': ProductName.objects.all().select_related('product_type'),
        'species_list': WoodSpecies.objects.all().order_by('name'),
        'grades_list': QualityGrade.objects.all().order_by('code'),
        'lumber_dims': LumberDimension.objects.all().order_by('thickness', 'width', 'length'),
        'unit_dims': UnitDimension.objects.all().order_by('length', 'width', 'height'),
        'delete_url': 'lumber_track:productitem_delete',
    }
    return render(request, 'lumber_track/directory_productitem.html', context)


def productitem_create(request):
    if request.method == 'POST':
        product_name_id = request.POST.get('product_name')
        species_id = request.POST.get('species')
        grade_id = request.POST.get('grade')
        lumber_dim_id = request.POST.get('lumber_dim')
        unit_dim_id = request.POST.get('unit_dim')
        if product_name_id and species_id and grade_id:
            ProductItem.objects.create(
                product_name_id=product_name_id,
                species_id=species_id,
                grade_id=grade_id,
                lumber_dim_id=lumber_dim_id or None,
                unit_dim_id=unit_dim_id or None,
                is_active=True
            )
    return redirect('lumber_track:productitem_list')


def productitem_delete(request, pk):
    item = get_object_or_404(ProductItem, pk=pk)
    item.delete()
    return redirect('lumber_track:productitem_list')


# ========== МЕСТА ХРАНЕНИЯ ==========
def storagelocation_list(request):
    items = StorageLocation.objects.all().order_by('name')
    context = {
        'title': 'Места хранения',
        'items': items,
        'create_url': '/directories/storagelocation/create/',
        'delete_url': 'lumber_track:storagelocation_delete',
    }
    return render(request, 'lumber_track/directory_table.html', context)


def storagelocation_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            StorageLocation.objects.create(name=name)
    return redirect('lumber_track:storagelocation_list')


def storagelocation_edit(request, pk):
    item = get_object_or_404(StorageLocation, pk=pk)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            item.name = name
            item.save()
    return redirect('lumber_track:storagelocation_list')


def storagelocation_delete(request, pk):
    item = get_object_or_404(StorageLocation, pk=pk)
    item.delete()
    return redirect('lumber_track:storagelocation_list')


def storagelocation_data(request, pk):
    item = get_object_or_404(StorageLocation, pk=pk)
    return JsonResponse({'name': item.name})


# ========== ДОКУМЕНТЫ ==========
def documents_page(request):
    return render(request, 'lumber_track/documents.html')


def document_journal(request, doc_type):
    doc_type_name = dict(Document.DOCUMENT_TYPES).get(doc_type, 'Документы')
    documents = Document.objects.filter(doc_type=doc_type).prefetch_related('items').order_by('-doc_date',
                                                                                              '-created_at')
    for doc in documents:
        doc.has_items = doc.items.exists()
    context = {
        'title': doc_type_name,
        'documents': documents,
        'doc_type': doc_type,
        'delete_url': 'lumber_track:document_delete',
    }
    return render(request, 'lumber_track/document_journal.html', context)


def document_delete(request, pk):
    doc = get_object_or_404(Document, pk=pk)
    doc_type = doc.doc_type
    doc_number = doc.doc_number
    if doc.items.exists():
        messages.error(request, f'❌ Невозможно удалить документ "{doc_number}", так как он содержит позиции.')
    else:
        doc.delete()
        messages.success(request, f'✅ Документ "{doc_number}" успешно удален.')
    if doc_type == 1:
        return redirect('lumber_track:document_initial_journal')
    elif doc_type == 2:
        return redirect('lumber_track:document_income_journal')
    else:
        return redirect('lumber_track:document_outcome_journal')


def document_create(request, doc_type):
    doc_type_name = dict(Document.DOCUMENT_TYPES).get(doc_type, 'Документ')

    if request.method == 'POST':
        doc_number = request.POST.get('doc_number', '').strip()
        if Document.objects.filter(doc_type=doc_type, doc_number=doc_number).exists():
            messages.error(request, f'Документ с номером "{doc_number}" для данного типа документа уже существует!')
            context = {
                'doc_type': doc_type,
                'doc_type_name': doc_type_name,
                'product_names': ProductName.objects.all().select_related('product_type'),
                'species_list': WoodSpecies.objects.all().order_by('name'),
                'grades_list': QualityGrade.objects.all().order_by('code'),
                'lumber_dims': LumberDimension.objects.all().order_by('thickness', 'width', 'length'),
                'unit_dims': UnitDimension.objects.all().order_by('length', 'width', 'height'),
                'locations': StorageLocation.objects.all().order_by('name'),
            }
            if doc_type == 3:
                available_stocks = get_available_stocks_with_details()
                context['available_stocks'] = available_stocks
                return render(request, 'lumber_track/document_outcome_form.html', context)
            return render(request, 'lumber_track/document_form.html', context)

        # Определяем location_id
        location_id = request.POST.get('location')
        if not location_id and doc_type == 1:
            location_id = 2
        if not location_id and doc_type == 3:
            location_id = 2

        doc = Document.objects.create(
            doc_type=doc_type,
            doc_number=doc_number,
            doc_date=request.POST.get('doc_date'),
            note=request.POST.get('note', ''),
            location_id=location_id
        )

        # Сохраняем to_location для расхода
        if doc_type == 3:
            to_location_id = request.POST.get('to_location')
            if to_location_id:
                doc.to_location_id = to_location_id
                doc.save()

        # Сохраняем позиции
        if doc_type == 3:
            stock_items = request.POST.getlist('stock_item[]')
            quantities = request.POST.getlist('quantity[]')
            product_names = request.POST.getlist('product_name[]')
            species_list = request.POST.getlist('species[]')
            grades_list = request.POST.getlist('grade[]')
            dimension_ids = request.POST.getlist('dimension_id[]')
            dimension_types = request.POST.getlist('dimension_type[]')

            for i in range(len(stock_items)):
                if not stock_items[i]:
                    continue
                quantity = int(quantities[i]) if i < len(quantities) and quantities[i] else 0
                if quantity == 0:
                    continue
                product_name_id = product_names[i] if i < len(product_names) else None
                species_id = species_list[i] if i < len(species_list) else None
                grade_id = grades_list[i] if i < len(grades_list) else None
                dimension_id = dimension_ids[i] if i < len(dimension_ids) else None
                dimension_type = dimension_types[i] if i < len(dimension_types) else None
                lumber_dim_id = None
                unit_dim_id = None
                if dimension_type == 'lumber':
                    lumber_dim_id = dimension_id
                elif dimension_type == 'unit':
                    unit_dim_id = dimension_id
                DocumentItem.objects.create(
                    document=doc,
                    product_name_id=product_name_id,
                    species_id=species_id,
                    grade_id=grade_id,
                    lumber_dim_id=lumber_dim_id,
                    unit_dim_id=unit_dim_id,
                    quantity=quantity
                )
        else:
            product_names = request.POST.getlist('product_name[]')
            species_list = request.POST.getlist('species[]')
            grades_list = request.POST.getlist('grade[]')
            dimension_ids = request.POST.getlist('dimension_id[]')
            dimension_types = request.POST.getlist('dimension_type[]')
            quantities = request.POST.getlist('quantity[]')

            for i in range(len(product_names)):
                if not product_names[i] or not species_list[i] or not grades_list[i]:
                    continue
                quantity = int(quantities[i]) if i < len(quantities) and quantities[i] else 0
                if quantity == 0:
                    continue
                lumber_dim_id = None
                unit_dim_id = None
                if i < len(dimension_ids) and dimension_ids[i]:
                    if i < len(dimension_types) and dimension_types[i] == 'lumber':
                        lumber_dim_id = dimension_ids[i]
                    elif i < len(dimension_types) and dimension_types[i] == 'unit':
                        unit_dim_id = dimension_ids[i]
                DocumentItem.objects.create(
                    document=doc,
                    product_name_id=product_names[i],
                    species_id=species_list[i],
                    grade_id=grades_list[i],
                    lumber_dim_id=lumber_dim_id,
                    unit_dim_id=unit_dim_id,
                    quantity=quantity
                )

        redirect_urls = {1: 'lumber_track:document_initial_journal', 2: 'lumber_track:document_income_journal',
                         3: 'lumber_track:document_outcome_journal'}
        return redirect(redirect_urls.get(doc_type, 'lumber_track:document_initial_journal'))

    # GET запрос
    context = {
        'doc_type': doc_type,
        'doc_type_name': doc_type_name,
        'product_names': ProductName.objects.all().select_related('product_type'),
        'species_list': WoodSpecies.objects.all().order_by('name'),
        'grades_list': QualityGrade.objects.all().order_by('code'),
        'lumber_dims': LumberDimension.objects.all().order_by('thickness', 'width', 'length'),
        'unit_dims': UnitDimension.objects.all().order_by('length', 'width', 'height'),
        'locations': StorageLocation.objects.all().order_by('name'),
    }
    if doc_type == 3:
        available_stocks = get_available_stocks_with_details()
        context['available_stocks'] = available_stocks
        return render(request, 'lumber_track/document_outcome_form.html', context)
    else:
        return render(request, 'lumber_track/document_form.html', context)


def document_edit(request, pk):
    doc = get_object_or_404(Document, pk=pk)
    doc_type = doc.doc_type
    doc_type_name = dict(Document.DOCUMENT_TYPES).get(doc_type, 'Документ')

    original_note = doc.note
    to_location_name = None
    to_location_id = doc.to_location_id

    # Для старых документов, где to_location было в примечании
    if doc_type == 3 and not doc.to_location_id and doc.note:
        match = re.search(r'Перемещено в: (.+)', doc.note)
        if match:
            to_location_name = match.group(1).strip()
            original_note = re.sub(r'\n?Перемещено в: .+', '', doc.note).strip()

    if request.method == 'POST':
        doc.doc_number = request.POST.get('doc_number')
        doc.doc_date = request.POST.get('doc_date')

        if doc_type == 3:
            to_location_id = request.POST.get('to_location')
            if to_location_id:
                doc.to_location_id = to_location_id
            doc.note = request.POST.get('note', '')
        else:
            doc.note = request.POST.get('note', '')
        doc.save()

        doc.items.all().delete()

        if doc_type == 3:
            stock_items = request.POST.getlist('stock_item[]')
            quantities = request.POST.getlist('quantity[]')
            product_names = request.POST.getlist('product_name[]')
            species_list = request.POST.getlist('species[]')
            grades_list = request.POST.getlist('grade[]')
            dimension_ids = request.POST.getlist('dimension_id[]')
            dimension_types = request.POST.getlist('dimension_type[]')

            for i in range(len(stock_items)):
                if not stock_items[i]:
                    continue
                quantity = int(quantities[i]) if i < len(quantities) and quantities[i] else 0
                if quantity == 0:
                    continue
                product_name_id = product_names[i] if i < len(product_names) else None
                species_id = species_list[i] if i < len(species_list) else None
                grade_id = grades_list[i] if i < len(grades_list) else None
                dimension_id = dimension_ids[i] if i < len(dimension_ids) else None
                dimension_type = dimension_types[i] if i < len(dimension_types) else None
                lumber_dim_id = None
                unit_dim_id = None
                if dimension_type == 'lumber':
                    lumber_dim_id = dimension_id
                elif dimension_type == 'unit':
                    unit_dim_id = dimension_id
                DocumentItem.objects.create(
                    document=doc,
                    product_name_id=product_name_id,
                    species_id=species_id,
                    grade_id=grade_id,
                    lumber_dim_id=lumber_dim_id,
                    unit_dim_id=unit_dim_id,
                    quantity=quantity
                )
        else:
            product_names = request.POST.getlist('product_name[]')
            species_list = request.POST.getlist('species[]')
            grades_list = request.POST.getlist('grade[]')
            dimension_ids = request.POST.getlist('dimension_id[]')
            dimension_types = request.POST.getlist('dimension_type[]')
            quantities = request.POST.getlist('quantity[]')

            for i in range(len(product_names)):
                if not product_names[i] or not species_list[i] or not grades_list[i]:
                    continue
                quantity = int(quantities[i]) if i < len(quantities) and quantities[i] else 0
                if quantity == 0:
                    continue
                lumber_dim_id = None
                unit_dim_id = None
                if i < len(dimension_ids) and dimension_ids[i]:
                    if i < len(dimension_types) and dimension_types[i] == 'lumber':
                        lumber_dim_id = dimension_ids[i]
                    elif i < len(dimension_types) and dimension_types[i] == 'unit':
                        unit_dim_id = dimension_ids[i]
                DocumentItem.objects.create(
                    document=doc,
                    product_name_id=product_names[i],
                    species_id=species_list[i],
                    grade_id=grades_list[i],
                    lumber_dim_id=lumber_dim_id,
                    unit_dim_id=unit_dim_id,
                    quantity=quantity
                )

        redirect_urls = {1: 'lumber_track:document_initial_journal', 2: 'lumber_track:document_income_journal',
                         3: 'lumber_track:document_outcome_journal'}
        return redirect(redirect_urls.get(doc_type, 'lumber_track:document_initial_journal'))

    context = {
        'doc': doc,
        'doc_type': doc_type,
        'doc_type_name': doc_type_name,
        'original_note': original_note,
        'to_location_id': to_location_id,
        'to_location_name': to_location_name,
        'product_names': ProductName.objects.all().select_related('product_type'),
        'species_list': WoodSpecies.objects.all().order_by('name'),
        'grades_list': QualityGrade.objects.all().order_by('code'),
        'lumber_dims': LumberDimension.objects.all().order_by('thickness', 'width', 'length'),
        'unit_dims': UnitDimension.objects.all().order_by('length', 'width', 'height'),
        'locations': StorageLocation.objects.all().order_by('name'),
        'items': doc.items.all().select_related('product_name', 'species', 'grade', 'lumber_dim', 'unit_dim'),
    }
    if doc_type == 3:
        available_stocks = get_available_stocks_with_details(location_id=2, exclude_document_id=doc.id)
        context['available_stocks'] = available_stocks
        return render(request, 'lumber_track/document_outcome_edit.html', context)
    else:
        return render(request, 'lumber_track/document_edit.html', context)


# ========== API ==========
@csrf_exempt
@require_http_methods(["POST"])
def api_add_productname(request):
    try:
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        product_type_id = data.get('product_type_id')
        if not name:
            return JsonResponse({'error': 'Название не может быть пустым'}, status=400)
        if not product_type_id:
            product_type, _ = ProductType.objects.get_or_create(name="Погонаж")
            product_type_id = product_type.id
        obj, created = ProductName.objects.get_or_create(
            name=name,
            defaults={'product_type_id': product_type_id}
        )
        return JsonResponse({
            'id': obj.id,
            'text': f"{obj.name} ({obj.product_type.name})",
            'created': created,
            'product_type_id': obj.product_type.id,
            'product_type_name': obj.product_type.name
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
@require_http_methods(["POST"])
def api_add_woodspecies(request):
    try:
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        if not name:
            return JsonResponse({'error': 'Название не может быть пустым'}, status=400)
        obj, created = WoodSpecies.objects.get_or_create(name=name)
        return JsonResponse({'id': obj.id, 'text': obj.name, 'created': created})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
@require_http_methods(["POST"])
def api_add_qualitygrade(request):
    try:
        data = json.loads(request.body)
        code = data.get('name', '').strip().upper()
        if not code:
            return JsonResponse({'error': 'Код не может быть пустым'}, status=400)
        obj, created = QualityGrade.objects.get_or_create(code=code)
        return JsonResponse({'id': obj.id, 'text': obj.code, 'created': created})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
@require_http_methods(["POST"])
def api_add_lumberdimension(request):
    try:
        data = json.loads(request.body)
        thickness = int(data.get('thickness', 0))
        width = int(data.get('width', 0))
        length = int(data.get('length', 0))
        if not all([thickness, width, length]):
            return JsonResponse({'error': 'Все размеры должны быть указаны'}, status=400)
        obj, created = LumberDimension.objects.get_or_create(
            thickness=thickness,
            width=width,
            length=length
        )
        return JsonResponse({
            'id': obj.id,
            'text': f"{thickness}-{width}-{length} мм ({obj.volume_m3:.6f} м³)",
            'created': created,
            'volume': obj.volume_m3,
            'area': obj.area_m2
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
@require_http_methods(["POST"])
def api_add_unitdimension(request):
    try:
        data = json.loads(request.body)
        length = int(data.get('length', 0))
        width = int(data.get('width', 0))
        height = int(data.get('height', 0))
        if not all([length, width, height]):
            return JsonResponse({'error': 'Все размеры должны быть указаны'}, status=400)
        obj, created = UnitDimension.objects.get_or_create(
            length=length,
            width=width,
            height=height
        )
        return JsonResponse({
            'id': obj.id,
            'text': f"{length}-{width}-{height} мм",
            'created': created
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


def api_get_lumberdim_data(request, pk):
    dim = get_object_or_404(LumberDimension, pk=pk)
    return JsonResponse({
        'volume': dim.volume_m3,
        'area': dim.area_m2
    })


@csrf_exempt
def api_add_dimension(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        thickness = int(data.get('thickness'))
        width = int(data.get('width'))
        length = int(data.get('length'))
        obj, created = LumberDimension.objects.get_or_create(
            thickness=thickness,
            width=width,
            length=length
        )
        return JsonResponse({
            'id': obj.id,
            'volume': obj.volume_m3,
            'area': obj.area_m2
        })
    return JsonResponse({'error': 'Ошибка'}, status=400)


def api_search_productname(request):
    term = request.GET.get('term', '')
    items = ProductName.objects.filter(name__icontains=term).select_related('product_type')[:20]
    results = [{'id': item.id, 'text': f"{item.name} ({item.product_type.name})"} for item in items]
    return JsonResponse({'results': results})


def api_search_woodspecies(request):
    term = request.GET.get('term', '')
    items = WoodSpecies.objects.filter(name__icontains=term)[:20]
    results = [{'id': item.id, 'text': item.name} for item in items]
    return JsonResponse({'results': results})


def api_search_qualitygrade(request):
    term = request.GET.get('term', '')
    items = QualityGrade.objects.filter(code__icontains=term)[:20]
    results = [{'id': item.id, 'text': item.code} for item in items]
    return JsonResponse({'results': results})


def api_search_lumberdim(request):
    term = request.GET.get('term', '')
    items = LumberDimension.objects.filter(
        models.Q(thickness__icontains=term) |
        models.Q(width__icontains=term) |
        models.Q(length__icontains=term)
    )[:20]
    results = [{'id': item.id, 'text': f"{item.thickness}-{item.width}-{item.length} мм ({item.volume_m3:.6f} м³)"} for
               item in items]
    return JsonResponse({'results': results})


def api_search_unitdim(request):
    term = request.GET.get('term', '')
    items = UnitDimension.objects.filter(
        models.Q(length__icontains=term) |
        models.Q(width__icontains=term) |
        models.Q(height__icontains=term)
    )[:20]
    results = [{'id': item.id, 'text': f"{item.length}-{item.width}-{item.height} мм"} for item in items]
    return JsonResponse({'results': results})


# ========== ОТЧЕТЫ ==========
def reports_page(request):
    return render(request, 'lumber_track/reports.html')


def report_income(request):
    if request.GET.get('date_from') and request.GET.get('date_to'):
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        return redirect(f'/reports/income/result/?date_from={date_from}&date_to={date_to}')
    return render(request, 'lumber_track/report_form.html', {
        'title': 'Поступление от ЦСИ',
        'url_name': 'lumber_track:report_income_result'
    })


def report_income_result(request):
    """Результат отчета по поступлению с группировкой по дням"""
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if not date_from or not date_to:
        return redirect('lumber_track:report_income')

    from datetime import datetime
    from collections import defaultdict

    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()

    # Получаем данные из приходных документов
    items = DocumentItem.objects.filter(
        document__doc_type=2,
        document__doc_date__gte=date_from_obj,
        document__doc_date__lte=date_to_obj
    ).select_related('document', 'product_name', 'species', 'grade', 'lumber_dim', 'unit_dim').order_by(
        'document__doc_date')

    # Группировка по дням
    daily_data = defaultdict(lambda: {
        'items': [],
        'total_quantity': 0,
        'total_volume': 0,
        'total_area': 0,
    })

    for item in items:
        doc_date = item.document.doc_date

        # Агрегируем по товарам внутри дня
        key = f"{item.product_name_id}_{item.species_id}_{item.grade_id}_{item.lumber_dim_id}_{item.unit_dim_id}"

        # Ищем существующий товар в этом дне
        existing_item = None
        for existing in daily_data[doc_date]['items']:
            if existing.get('key') == key:
                existing_item = existing
                break

        if existing_item:
            existing_item['total_quantity'] += item.quantity
            existing_item['total_volume'] += item.volume_m3
            existing_item['total_area'] += item.area_m2
        else:
            daily_data[doc_date]['items'].append({
                'key': key,
                'product_name': item.product_name.name,
                'species': item.species.name,
                'grade': item.grade.code,
                'dimension_display': item.dimension_display,
                'total_quantity': item.quantity,
                'total_volume': item.volume_m3,
                'total_area': item.area_m2,
            })

        # Суммируем итоги дня
        daily_data[doc_date]['total_quantity'] += item.quantity
        daily_data[doc_date]['total_volume'] += item.volume_m3
        daily_data[doc_date]['total_area'] += item.area_m2

    # Преобразуем в список для шаблона
    report_data = []
    for doc_date in sorted(daily_data.keys()):
        # Заголовок даты
        report_data.append({
            'is_date_header': True,
            'date': doc_date,
        })
        # Детальные строки
        for item in daily_data[doc_date]['items']:
            report_data.append({
                'is_date_header': False,
                'is_day_total': False,
                'product_name': item['product_name'],
                'species': item['species'],
                'grade': item['grade'],
                'dimension_display': item['dimension_display'],
                'total_quantity': item['total_quantity'],
                'total_volume': item['total_volume'],
                'total_area': item['total_area'],
            })
        # Итоги за день
        report_data.append({
            'is_date_header': False,
            'is_day_total': True,
            'total_quantity': daily_data[doc_date]['total_quantity'],
            'total_volume': daily_data[doc_date]['total_volume'],
            'total_area': daily_data[doc_date]['total_area'],
        })
    # Общие итоги за весь период
    grand_total_quantity = sum(item['total_quantity'] for item in report_data if not item.get('is_date_header'))
    grand_total_volume = sum(item['total_volume'] for item in report_data if not item.get('is_date_header'))
    grand_total_area = sum(item['total_area'] for item in report_data if not item.get('is_date_header'))

    context = {
        'title': 'Поступление от ЦСИ',
        'date_from': date_from_obj,
        'date_to': date_to_obj,
        'report_data': report_data,
        'grand_total_quantity': grand_total_quantity,
        'grand_total_volume': grand_total_volume,
        'grand_total_area': grand_total_area,
        'url_name': 'lumber_track:report_income_result',
    }

    if request.GET.get('export') == 'excel':
        return export_income_to_excel(report_data, context['title'], date_from, date_to,
                                      grand_total_quantity, grand_total_volume, grand_total_area)

    return render(request, 'lumber_track/report_result.html', context)
def report_to_stock(request):
    if request.GET.get('date_from') and request.GET.get('date_to'):
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        return redirect(f'/reports/to-stock/result/?date_from={date_from}&date_to={date_to}')
    return render(request, 'lumber_track/report_form.html', {
        'title': 'Склад-магазин Стрелка"',
        'url_name': 'lumber_track:report_to_stock_result'
    })


def report_to_stock_result(request):
    """Результат отчета по продукции на склад с группировкой по дням"""
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if not date_from or not date_to:
        return redirect('lumber_track:report_to_stock')

    from datetime import datetime
    from collections import defaultdict

    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()

    # Получаем данные из расходных документов (doc_type=3, to_location=4 - розничный склад)
    items = DocumentItem.objects.filter(
        document__doc_type=3,
        document__to_location_id=4,
        document__doc_date__gte=date_from_obj,
        document__doc_date__lte=date_to_obj
    ).select_related('document', 'product_name', 'species', 'grade', 'lumber_dim', 'unit_dim').order_by(
        'document__doc_date')

    # Группировка по дням
    daily_data = defaultdict(lambda: {
        'items': [],
        'total_quantity': 0,
        'total_volume': 0,
        'total_area': 0,
    })

    for item in items:
        doc_date = item.document.doc_date

        # Агрегируем по товарам внутри дня
        key = f"{item.product_name_id}_{item.species_id}_{item.grade_id}_{item.lumber_dim_id}_{item.unit_dim_id}"

        # Ищем существующий товар в этом дне
        existing_item = None
        for existing in daily_data[doc_date]['items']:
            if existing.get('key') == key:
                existing_item = existing
                break

        if existing_item:
            existing_item['total_quantity'] += item.quantity
            existing_item['total_volume'] += item.volume_m3
            existing_item['total_area'] += item.area_m2
        else:
            daily_data[doc_date]['items'].append({
                'key': key,
                'product_name': item.product_name.name,
                'species': item.species.name,
                'grade': item.grade.code,
                'dimension_display': item.dimension_display,
                'total_quantity': item.quantity,
                'total_volume': item.volume_m3,
                'total_area': item.area_m2,
            })

        # Суммируем итоги дня
        daily_data[doc_date]['total_quantity'] += item.quantity
        daily_data[doc_date]['total_volume'] += item.volume_m3
        daily_data[doc_date]['total_area'] += item.area_m2

    # Преобразуем в список для шаблона
    report_data = []
    for doc_date in sorted(daily_data.keys()):
        # Заголовок даты
        report_data.append({
            'is_date_header': True,
            'date': doc_date,
        })
        # Детальные строки
        for item in daily_data[doc_date]['items']:
            report_data.append({
                'is_date_header': False,
                'is_day_total': False,
                'product_name': item['product_name'],
                'species': item['species'],
                'grade': item['grade'],
                'dimension_display': item['dimension_display'],
                'total_quantity': item['total_quantity'],
                'total_volume': item['total_volume'],
                'total_area': item['total_area'],
            })
        # Итоги за день
        report_data.append({
            'is_date_header': False,
            'is_day_total': True,
            'total_quantity': daily_data[doc_date]['total_quantity'],
            'total_volume': daily_data[doc_date]['total_volume'],
            'total_area': daily_data[doc_date]['total_area'],
        })

    # Общие итоги за весь период
    grand_total_quantity = sum(item['total_quantity'] for item in report_data if
                               not item.get('is_date_header') and not item.get('is_day_total'))
    grand_total_volume = sum(
        item['total_volume'] for item in report_data if not item.get('is_date_header') and not item.get('is_day_total'))
    grand_total_area = sum(
        item['total_area'] for item in report_data if not item.get('is_date_header') and not item.get('is_day_total'))

    context = {
        'title': 'Склад-магазин Стрелка',
        'date_from': date_from_obj,
        'date_to': date_to_obj,
        'report_data': report_data,
        'grand_total_quantity': grand_total_quantity,
        'grand_total_volume': grand_total_volume,
        'grand_total_area': grand_total_area,
        'url_name': 'lumber_track:report_to_stock_result',
    }

    if request.GET.get('export') == 'excel':
        return export_to_stock_excel(report_data, context['title'], date_from, date_to,
                                     grand_total_quantity, grand_total_volume, grand_total_area)

    return render(request, 'lumber_track/report_result.html', context)

def report_to_shop(request):
    if request.GET.get('date_from') and request.GET.get('date_to'):
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        return redirect(f'/reports/to-shop/result/?date_from={date_from}&date_to={date_to}')
    return render(request, 'lumber_track/report_form.html', {
        'title': 'Отчет "В магазин"',
        'url_name': 'lumber_track:report_to_shop_result'
    })


def report_to_shop_result(request):
    """Результат отчета по продукции в магазин с группировкой по дням"""
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if not date_from or not date_to:
        return redirect('lumber_track:report_to_shop')

    from datetime import datetime
    from collections import defaultdict

    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()

    # Получаем данные из расходных документов (doc_type=3, to_location=3 - ТД Ангара)
    items = DocumentItem.objects.filter(
        document__doc_type=3,
        document__to_location_id=3,
        document__doc_date__gte=date_from_obj,
        document__doc_date__lte=date_to_obj
    ).select_related('document', 'product_name', 'species', 'grade', 'lumber_dim', 'unit_dim').order_by(
        'document__doc_date')

    # Группировка по дням
    daily_data = defaultdict(lambda: {
        'items': [],
        'total_quantity': 0,
        'total_volume': 0,
        'total_area': 0,
    })

    for item in items:
        doc_date = item.document.doc_date

        # Агрегируем по товарам внутри дня
        key = f"{item.product_name_id}_{item.species_id}_{item.grade_id}_{item.lumber_dim_id}_{item.unit_dim_id}"

        # Ищем существующий товар в этом дне
        existing_item = None
        for existing in daily_data[doc_date]['items']:
            if existing.get('key') == key:
                existing_item = existing
                break

        if existing_item:
            existing_item['total_quantity'] += item.quantity
            existing_item['total_volume'] += item.volume_m3
            existing_item['total_area'] += item.area_m2
        else:
            daily_data[doc_date]['items'].append({
                'key': key,
                'product_name': item.product_name.name,
                'species': item.species.name,
                'grade': item.grade.code,
                'dimension_display': item.dimension_display,
                'total_quantity': item.quantity,
                'total_volume': item.volume_m3,
                'total_area': item.area_m2,
            })

        # Суммируем итоги дня
        daily_data[doc_date]['total_quantity'] += item.quantity
        daily_data[doc_date]['total_volume'] += item.volume_m3
        daily_data[doc_date]['total_area'] += item.area_m2

    # Преобразуем в список для шаблона
    report_data = []
    for doc_date in sorted(daily_data.keys()):
        # Заголовок даты
        report_data.append({
            'is_date_header': True,
            'date': doc_date,
        })
        # Детальные строки
        for item in daily_data[doc_date]['items']:
            report_data.append({
                'is_date_header': False,
                'is_day_total': False,
                'product_name': item['product_name'],
                'species': item['species'],
                'grade': item['grade'],
                'dimension_display': item['dimension_display'],
                'total_quantity': item['total_quantity'],
                'total_volume': item['total_volume'],
                'total_area': item['total_area'],
            })
        # Итоги за день
        report_data.append({
            'is_date_header': False,
            'is_day_total': True,
            'total_quantity': daily_data[doc_date]['total_quantity'],
            'total_volume': daily_data[doc_date]['total_volume'],
            'total_area': daily_data[doc_date]['total_area'],
        })

    # Общие итоги за весь период
    grand_total_quantity = sum(item['total_quantity'] for item in report_data if
                               not item.get('is_date_header') and not item.get('is_day_total'))
    grand_total_volume = sum(
        item['total_volume'] for item in report_data if not item.get('is_date_header') and not item.get('is_day_total'))
    grand_total_area = sum(
        item['total_area'] for item in report_data if not item.get('is_date_header') and not item.get('is_day_total'))

    context = {
        'title': 'Магазин в Красноярске"',
        'date_from': date_from_obj,
        'date_to': date_to_obj,
        'report_data': report_data,
        'grand_total_quantity': grand_total_quantity,
        'grand_total_volume': grand_total_volume,
        'grand_total_area': grand_total_area,
        'url_name': 'lumber_track:report_to_shop_result',
    }

    if request.GET.get('export') == 'excel':
        return export_to_shop_excel(report_data, context['title'], date_from, date_to,
                                    grand_total_quantity, grand_total_volume, grand_total_area)

    return render(request, 'lumber_track/report_result.html', context)

# lumber_track/views.py - добавьте после других функций отчетов

def report_movement(request):
    """Отчет по движению продукции (остатки, приход, расход)"""
    if request.GET.get('date_from') and request.GET.get('date_to'):
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        return redirect(f'/reports/movement/result/?date_from={date_from}&date_to={date_to}')
    return render(request, 'lumber_track/report_form.html', {
        'title': 'Отчет по движению продукции',
        'url_name': 'lumber_track:report_movement_result'
    })


# lumber_track/views.py - замените функцию report_movement_result

def report_movement_result(request):
    """Результат отчета по движению продукции"""
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if not date_from or not date_to:
        return redirect('lumber_track:report_movement')

    from django.db.models import Sum, Q
    from datetime import datetime

    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()

    # Собираем данные по каждой позиции
    report_data = {}

    # Получаем все позиции из документов за период и до периода
    all_items = DocumentItem.objects.select_related(
        'product_name', 'species', 'grade', 'lumber_dim', 'unit_dim'
    ).filter(
        document__doc_date__lte=date_to_obj
    )

    for item in all_items:
        # Определяем размер
        if item.lumber_dim:
            dimension_display = f"{item.lumber_dim.thickness}-{item.lumber_dim.width}-{item.lumber_dim.length} мм"
            dimension_type = 'lumber'
            dimension_id = item.lumber_dim.id
        elif item.unit_dim:
            dimension_display = f"{item.unit_dim.length}-{item.unit_dim.width}-{item.unit_dim.height} мм"
            dimension_type = 'unit'
            dimension_id = item.unit_dim.id
        else:
            dimension_display = "—"
            dimension_type = None
            dimension_id = None

        key = f"{item.product_name_id}_{item.species_id}_{item.grade_id}_{dimension_id}_{dimension_type}"

        if key not in report_data:
            report_data[key] = {
                'product_name': item.product_name.name,
                'product_name_id': item.product_name_id,
                'species': item.species.name,
                'species_id': item.species_id,
                'grade': item.grade.code,
                'grade_id': item.grade_id,
                'dimension_display': dimension_display,
                'dimension_type': dimension_type,
                'dimension_id': dimension_id,
                'initial_quantity': 0,
                'income_quantity': 0,
                'expense_quantity': 0,
            }

    # Считаем данные по каждому ключу
    for key, data in report_data.items():
        product_name_id = data['product_name_id']
        species_id = data['species_id']
        grade_id = data['grade_id']
        dimension_id = data['dimension_id']
        dimension_type = data['dimension_type']

        # Остаток на начало (документы с датой < date_from)
        # Учитываем: начальные остатки (doc_type=1), приходы на склад (doc_type=2, location_id=2)
        # и перемещения на склад (doc_type=3, to_location_id=2)
        initial_filter = Q(document__doc_date__lt=date_from_obj) & (
                (Q(document__doc_type=1) & Q(document__location_id=2)) |
                (Q(document__doc_type=2) & Q(document__location_id=2)) |
                (Q(document__doc_type=3) & Q(document__to_location_id=2))
        )

        # Приход за период
        income_filter = Q(document__doc_date__gte=date_from_obj, document__doc_date__lte=date_to_obj) & (
                (Q(document__doc_type=1) & Q(document__location_id=2)) |
                (Q(document__doc_type=2) & Q(document__location_id=2)) |
                (Q(document__doc_type=3) & Q(document__to_location_id=2))
        )

        # Расход за период (отгрузки со склада)
        expense_filter = Q(document__doc_date__gte=date_from_obj, document__doc_date__lte=date_to_obj) & (
                Q(document__doc_type=3) & ~Q(document__to_location_id=2)
        )

        # Базовый фильтр по товару
        base_filter = Q(
            product_name_id=product_name_id,
            species_id=species_id,
            grade_id=grade_id
        )

        # Фильтрация по размерам
        if dimension_type == 'lumber' and dimension_id:
            base_filter &= Q(lumber_dim_id=dimension_id)
        elif dimension_type == 'unit' and dimension_id:
            base_filter &= Q(unit_dim_id=dimension_id)
        else:
            base_filter &= Q(lumber_dim__isnull=True, unit_dim__isnull=True)

        # Получаем суммы
        initial_qs = DocumentItem.objects.filter(initial_filter & base_filter)
        income_qs = DocumentItem.objects.filter(income_filter & base_filter)
        expense_qs = DocumentItem.objects.filter(expense_filter & base_filter)

        initial_total = initial_qs.aggregate(total=Sum('quantity'))['total'] or 0
        income_total = income_qs.aggregate(total=Sum('quantity'))['total'] or 0
        expense_total = expense_qs.aggregate(total=Sum('quantity'))['total'] or 0

        data['initial_quantity'] = initial_total
        data['income_quantity'] = income_total
        data['expense_quantity'] = expense_total
        data['ending_quantity'] = initial_total + income_total - expense_total

    # Фильтруем позиции с ненулевыми значениями
    report_list = [data for data in report_data.values()
                   if data['initial_quantity'] != 0 or data['income_quantity'] != 0 or data['expense_quantity'] != 0 or
                   data['ending_quantity'] != 0]

    # Сортируем
    report_list.sort(key=lambda x: (x['product_name'], x['dimension_display'], x['species'], x['grade']))

    # Группируем по наименованию
    grouped_data = {}
    for item in report_list:
        name = item['product_name']
        if name not in grouped_data:
            grouped_data[name] = {
                'product_name': name,
                'items': [],
                'total_initial': 0,
                'total_income': 0,
                'total_expense': 0,
                'total_ending': 0,
            }
        grouped_data[name]['items'].append(item)
        grouped_data[name]['total_initial'] += item['initial_quantity']
        grouped_data[name]['total_income'] += item['income_quantity']
        grouped_data[name]['total_expense'] += item['expense_quantity']
        grouped_data[name]['total_ending'] += item['ending_quantity']

    # Общие итоги
    grand_total_initial = sum(data['total_initial'] for data in grouped_data.values())
    grand_total_income = sum(data['total_income'] for data in grouped_data.values())
    grand_total_expense = sum(data['total_expense'] for data in grouped_data.values())
    grand_total_ending = sum(data['total_ending'] for data in grouped_data.values())

    context = {
        'title': 'Отчет по движению продукции',
        'date_from': date_from_obj,
        'date_to': date_to_obj,
        'grouped_data': grouped_data.values(),
        'grand_total_initial': grand_total_initial,
        'grand_total_income': grand_total_income,
        'grand_total_expense': grand_total_expense,
        'grand_total_ending': grand_total_ending,
        'url_name': 'lumber_track:report_movement_result',
    }

    if request.GET.get('export') == 'excel':
        return export_movement_to_excel(grouped_data.values(), context['title'], date_from, date_to,
                                        grand_total_initial, grand_total_income, grand_total_expense,
                                        grand_total_ending)

    return render(request, 'lumber_track/report_movement.html', context)


def export_movement_to_excel(grouped_data, title, date_from, date_to,
                             grand_total_initial, grand_total_income, grand_total_expense, grand_total_ending):
    """Экспорт отчета по движению в Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Движение продукции"

    # Заголовки
    headers = ['Наименование', 'Размер', 'Порода', 'Сорт', 'Остаток на начало', 'Приход', 'Расход', 'Остаток на конец']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    row = 2
    for group in grouped_data:
        # Строки с деталями
        for item in group['items']:
            ws.cell(row=row, column=1, value=item['product_name'])
            ws.cell(row=row, column=2, value=item['dimension_display'])
            ws.cell(row=row, column=3, value=item['species'])
            ws.cell(row=row, column=4, value=item['grade'])
            ws.cell(row=row, column=5, value=float(item['initial_quantity']))
            ws.cell(row=row, column=6, value=float(item['income_quantity']))
            ws.cell(row=row, column=7, value=float(item['expense_quantity']))
            ws.cell(row=row, column=8, value=float(item['ending_quantity']))
            row += 1

        # Строка итогов по группе
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col, value=f"Итого: {group['product_name']}" if col == 1 else "")
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        ws.cell(row=row, column=5, value=float(group['total_initial']))
        ws.cell(row=row, column=6, value=float(group['total_income']))
        ws.cell(row=row, column=7, value=float(group['total_expense']))
        ws.cell(row=row, column=8, value=float(group['total_ending']))
        row += 1

    # Общие итоги
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col, value="ВСЕГО:" if col == 1 else "")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws.cell(row=row, column=5, value=float(grand_total_initial))
    ws.cell(row=row, column=6, value=float(grand_total_income))
    ws.cell(row=row, column=7, value=float(grand_total_expense))
    ws.cell(row=row, column=8, value=float(grand_total_ending))

    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{title}_{date_from}_{date_to}.xlsx"'
    wb.save(response)
    return response


# lumber_track/views.py - добавьте

def report_category(request):
    """Отчет по категориям (сводный)"""
    if request.GET.get('date_from') and request.GET.get('date_to'):
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        return redirect(f'/reports/category/result/?date_from={date_from}&date_to={date_to}')
    return render(request, 'lumber_track/report_form.html', {
        'title': 'Сводный отчет по категориям',
        'url_name': 'lumber_track:report_category_result'
    })


def report_category_result(request):
    """Результат сводного отчета по категориям"""
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if not date_from or not date_to:
        return redirect('lumber_track:report_category')

    from django.db.models import Sum, Q
    from datetime import datetime

    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()

    # Получаем все уникальные позиции (наименование + размер)
    # Получаем все уникальные позиции только для погонажа
    all_items = DocumentItem.objects.select_related(
        'product_name', 'species', 'grade', 'lumber_dim', 'unit_dim'
    ).filter(
        document__doc_date__lte=date_to_obj,
        lumber_dim__isnull=False  # Только погонаж (есть lumber_dim)
    ).exclude(
        lumber_dim_id__isnull=True  # Исключаем штучные
    )

    # Собираем данные по каждой позиции
    report_data = {}
    categories = ['A', 'B', 'C', 'D']

    for item in all_items:
        # Определяем размер
        if item.lumber_dim:
            thickness = item.lumber_dim.thickness
            width = item.lumber_dim.width
            length = item.lumber_dim.length
            dimension_display = f"{thickness}×{width}×{length}"
            dimension_type = 'lumber'
            dimension_id = item.lumber_dim.id
        elif item.unit_dim:
            thickness = item.unit_dim.length
            width = item.unit_dim.width
            height = item.unit_dim.height
            dimension_display = f"{thickness}×{width}×{height}"
            dimension_type = 'unit'
            dimension_id = item.unit_dim.id
        else:
            thickness = width = length = 0
            dimension_display = "—"
            dimension_type = None
            dimension_id = None

        key = f"{item.product_name_id}_{dimension_id}_{dimension_type}"

        if key not in report_data:
            report_data[key] = {
                'product_name': item.product_name.name,
                'species': item.species.name,
                'thickness': thickness,
                'width': width,
                'length': length,
                'dimension_display': dimension_display,
                'dimension_type': dimension_type,
                'dimension_id': dimension_id,
                'initial': {cat: 0 for cat in categories},
                'initial_volume': {cat: 0 for cat in categories},
                'initial_area': {cat: 0 for cat in categories},
                'income_production': {cat: 0 for cat in categories},
                'income_production_volume': {cat: 0 for cat in categories},
                'income_production_area': {cat: 0 for cat in categories},
                'to_stock': {cat: 0 for cat in categories},
                'to_stock_volume': {cat: 0 for cat in categories},
                'to_stock_area': {cat: 0 for cat in categories},
                'to_shop': {cat: 0 for cat in categories},
                'to_shop_volume': {cat: 0 for cat in categories},
                'to_shop_area': {cat: 0 for cat in categories},
                'ending': {cat: 0 for cat in categories},
                'ending_volume': {cat: 0 for cat in categories},
                'ending_area': {cat: 0 for cat in categories},
            }

    # Собираем данные по каждому ключу
    for key, data in report_data.items():
        product_name = data['product_name']
        dimension_id = data['dimension_id']
        dimension_type = data['dimension_type']

        for grade_code in categories:
            grade = QualityGrade.objects.filter(code=grade_code).first()
            if not grade:
                continue

            grade_id = grade.id

            # 1. Остаток на начало
            initial_filter = Q(document__doc_date__lt=date_from_obj) & (
                    (Q(document__doc_type=1) & Q(document__location_id=2)) |
                    (Q(document__doc_type=2) & Q(document__location_id=2)) |
                    (Q(document__doc_type=3) & Q(document__to_location_id=2))
            ) & Q(product_name__name=product_name, grade_id=grade_id)

            # 2. Приход из цеха
            income_production_filter = Q(
                document__doc_date__gte=date_from_obj,
                document__doc_date__lte=date_to_obj,
                document__doc_type=2,
                document__location_id=2
            ) & Q(product_name__name=product_name, grade_id=grade_id)

            # 3. Передано на склад (Розничный склад, id=4)
            to_stock_filter = Q(
                document__doc_date__gte=date_from_obj,
                document__doc_date__lte=date_to_obj,
                document__doc_type=3,
                document__to_location_id=4
            ) & Q(product_name__name=product_name, grade_id=grade_id)

            # 4. Передано в магазин (ТД Ангара, id=3)
            to_shop_filter = Q(
                document__doc_date__gte=date_from_obj,
                document__doc_date__lte=date_to_obj,
                document__doc_type=3,
                document__to_location_id=3
            ) & Q(product_name__name=product_name, grade_id=grade_id)
            # Фильтрация по размерам
            if dimension_type == 'lumber' and dimension_id:
                initial_filter &= Q(lumber_dim_id=dimension_id)
                income_production_filter &= Q(lumber_dim_id=dimension_id)
                to_stock_filter &= Q(lumber_dim_id=dimension_id)
                to_shop_filter &= Q(lumber_dim_id=dimension_id)
            elif dimension_type == 'unit' and dimension_id:
                initial_filter &= Q(unit_dim_id=dimension_id)
                income_production_filter &= Q(unit_dim_id=dimension_id)
                to_stock_filter &= Q(unit_dim_id=dimension_id)
                to_shop_filter &= Q(unit_dim_id=dimension_id)

            initial_qs = DocumentItem.objects.filter(initial_filter)
            income_production_qs = DocumentItem.objects.filter(income_production_filter)
            to_stock_qs = DocumentItem.objects.filter(to_stock_filter)
            to_shop_qs = DocumentItem.objects.filter(to_shop_filter)

            initial_qty = initial_qs.aggregate(total=Sum('quantity'))['total'] or 0
            income_production_qty = income_production_qs.aggregate(total=Sum('quantity'))['total'] or 0
            to_stock_qty = to_stock_qs.aggregate(total=Sum('quantity'))['total'] or 0
            to_shop_qty = to_shop_qs.aggregate(total=Sum('quantity'))['total'] or 0

            # Объем и площадь для погонажа
            if dimension_type == 'lumber' and dimension_id:
                try:
                    lumber_dim = LumberDimension.objects.get(id=dimension_id)
                    volume_per_unit = lumber_dim.volume_m3
                    area_per_unit = lumber_dim.area_m2
                except LumberDimension.DoesNotExist:
                    volume_per_unit = 0
                    area_per_unit = 0
            else:
                volume_per_unit = 0
                area_per_unit = 0

            data['initial'][grade_code] = initial_qty
            data['initial_volume'][grade_code] = initial_qty * volume_per_unit
            data['initial_area'][grade_code] = initial_qty * area_per_unit

            data['income_production'][grade_code] = income_production_qty
            data['income_production_volume'][grade_code] = income_production_qty * volume_per_unit
            data['income_production_area'][grade_code] = income_production_qty * area_per_unit

            data['to_stock'][grade_code] = to_stock_qty
            data['to_stock_volume'][grade_code] = to_stock_qty * volume_per_unit
            data['to_stock_area'][grade_code] = to_stock_qty * area_per_unit

            data['to_shop'][grade_code] = to_shop_qty
            data['to_shop_volume'][grade_code] = to_shop_qty * volume_per_unit
            data['to_shop_area'][grade_code] = to_shop_qty * area_per_unit

            # Исправленная формула: начальный + приход - склад - магазин
            data['ending'][grade_code] = initial_qty + income_production_qty - to_stock_qty - to_shop_qty
            data['ending_volume'][grade_code] = (initial_qty + income_production_qty - to_stock_qty - to_shop_qty) * volume_per_unit
            data['ending_area'][grade_code] = (initial_qty + income_production_qty - to_stock_qty - to_shop_qty) * area_per_unit
    # Фильтруем позиции с данными
    report_list = []
    for key, data in report_data.items():
        has_data = False
        for cat in categories:
            if (data['initial'][cat] != 0 or
                data['income_production'][cat] != 0 or
                data['to_stock'][cat] != 0 or
                data['to_shop'][cat] != 0 or
                data['ending'][cat] != 0):
                has_data = True
                break
        if has_data:
            report_list.append(data)

    report_list.sort(key=lambda x: x['product_name'])

    # Группировка по наименованию с итогами
    # Группировка по породе, а внутри по наименованию
    grouped_by_species = {}
    for item in report_list:
        species = item['species']  # теперь это поле есть
        product_name = item['product_name']

        if species not in grouped_by_species:
            grouped_by_species[species] = {}

        if product_name not in grouped_by_species[species]:
            grouped_by_species[species][product_name] = {
                'product_name': product_name,
                'species': species,
                'items': [],
                'total_initial': {cat: 0 for cat in categories},
                'total_initial_volume': {cat: 0 for cat in categories},
                'total_initial_area': {cat: 0 for cat in categories},
                'total_income_production': {cat: 0 for cat in categories},
                'total_income_production_volume': {cat: 0 for cat in categories},
                'total_income_production_area': {cat: 0 for cat in categories},
                'total_to_stock': {cat: 0 for cat in categories},
                'total_to_stock_volume': {cat: 0 for cat in categories},
                'total_to_stock_area': {cat: 0 for cat in categories},
                'total_to_shop': {cat: 0 for cat in categories},
                'total_to_shop_volume': {cat: 0 for cat in categories},
                'total_to_shop_area': {cat: 0 for cat in categories},
                'total_ending': {cat: 0 for cat in categories},
                'total_ending_volume': {cat: 0 for cat in categories},
                'total_ending_area': {cat: 0 for cat in categories},
            }

        # Добавляем детальную запись
        grouped_by_species[species][product_name]['items'].append(item)

        for cat in categories:
            grouped_by_species[species][product_name]['total_initial'][cat] += item['initial'][cat]
            grouped_by_species[species][product_name]['total_initial_volume'][cat] += item['initial_volume'][cat]
            grouped_by_species[species][product_name]['total_initial_area'][cat] += item['initial_area'][cat]
            grouped_by_species[species][product_name]['total_income_production'][cat] += item['income_production'][cat]
            grouped_by_species[species][product_name]['total_income_production_volume'][cat] += \
            item['income_production_volume'][cat]
            grouped_by_species[species][product_name]['total_income_production_area'][cat] += \
            item['income_production_area'][cat]
            grouped_by_species[species][product_name]['total_to_stock'][cat] += item['to_stock'][cat]
            grouped_by_species[species][product_name]['total_to_stock_volume'][cat] += item['to_stock_volume'][cat]
            grouped_by_species[species][product_name]['total_to_stock_area'][cat] += item['to_stock_area'][cat]
            grouped_by_species[species][product_name]['total_to_shop'][cat] += item['to_shop'][cat]
            grouped_by_species[species][product_name]['total_to_shop_volume'][cat] += item['to_shop_volume'][cat]
            grouped_by_species[species][product_name]['total_to_shop_area'][cat] += item['to_shop_area'][cat]
            grouped_by_species[species][product_name]['total_ending'][cat] += item['ending'][cat]
            grouped_by_species[species][product_name]['total_ending_volume'][cat] += item['ending_volume'][cat]
            grouped_by_species[species][product_name]['total_ending_area'][cat] += item['ending_area'][cat]
    # Добавляем в контекст новую структуру
    context = {
        'title': 'Сводный отчет по категориям',
        'date_from': date_from_obj,
        'date_to': date_to_obj,
        'grouped_by_species': grouped_by_species,  # <-- проверьте, что здесь grouped_by_species
        'categories': categories,
        'url_name': 'lumber_track:report_category_result',
    }

    if request.GET.get('export') == 'excel':
        return export_category_to_excel(grouped_by_species, categories, context['title'], date_from, date_to)

    return render(request, 'lumber_track/report_category.html', context)


def export_category_to_excel(grouped_by_species, categories, title, date_from, date_to):
    """Экспорт сводного отчета в Excel с группировкой по породам"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    subheader_font = Font(bold=True, size=10)
    subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    species_fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')
    total_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 1. Первая строка заголовков (основные секции)
    headers_row1 = ['Наименование', 'Размер, мм']
    sections = ['Остаток на начало', 'Поступление из цеха', 'Передано на склад', 'Передано в магазин',
                'Остаток на конец']
    for section in sections:
        headers_row1.append(section)  # Временно, потом объединим

    # Записываем первую строку
    col = 1
    ws.cell(row=1, column=col, value='Наименование')
    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
    col += 1

    ws.cell(row=1, column=col, value='Размер, мм')
    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col + 2)
    col += 3

    for section in sections:
        ws.cell(row=1, column=col, value=section)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 11)
        col += 12

    # 2. Вторая строка заголовков (типы показателей)
    col = 1
    col += 4  # пропускаем наименование и размеры
    for _ in range(5):  # 5 секций
        ws.cell(row=2, column=col, value='Кол-во (шт)')
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 3)
        col += 4
        ws.cell(row=2, column=col, value='Объем (м³)')
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 3)
        col += 4
        ws.cell(row=2, column=col, value='Площадь (м²)')
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 3)
        col += 4

    # 3. Третья строка заголовков (категории A,B,C,D)
    col = 5
    for _ in range(15):  # 5 секций × 3 показателя = 15 групп по 4 категории
        for cat in categories:
            ws.cell(row=3, column=col, value=cat)
            col += 1

    # Оформление заголовков
    for col in range(1, 65):
        for row in range(1, 4):
            cell = ws.cell(row=row, column=col)
            if row == 1:
                cell.font = header_font
                cell.fill = header_fill
            elif row == 2:
                cell.font = subheader_font
                cell.fill = subheader_fill
            elif row == 3:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 4. Данные
    current_row = 4
    for species, groups in grouped_by_species.items():
        # Заголовок породы
        cell = ws.cell(row=current_row, column=1, value=f"🌲 {species}")
        cell.font = Font(bold=True, size=11)
        cell.fill = species_fill
        for col in range(1, 65):
            ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1

        for group in groups.values():
            # Детальные строки
            for item in group['items']:
                # Размеры
                ws.cell(row=current_row, column=1, value=item['product_name'])
                ws.cell(row=current_row, column=2, value=item['thickness'] if item['thickness'] else '-')
                ws.cell(row=current_row, column=3, value=item['width'] if item['width'] else '-')
                ws.cell(row=current_row, column=4, value=item['length'] if item['length'] else '-')

                col = 5
                # Остаток на начало
                for cat in categories:
                    ws.cell(row=current_row, column=col,
                            value=item['initial'][cat] if item['initial'][cat] != 0 else '')
                    col += 1
                for cat in categories:
                    val = item['initial_volume'][cat]
                    ws.cell(row=current_row, column=col, value=round(val, 3) if val != 0 else '')
                    col += 1
                for cat in categories:
                    val = item['initial_area'][cat]
                    ws.cell(row=current_row, column=col, value=round(val, 2) if val != 0 else '')
                    col += 1
                # Поступление
                for cat in categories:
                    ws.cell(row=current_row, column=col,
                            value=item['income_production'][cat] if item['income_production'][cat] != 0 else '')
                    col += 1
                for cat in categories:
                    val = item['income_production_volume'][cat]
                    ws.cell(row=current_row, column=col, value=round(val, 3) if val != 0 else '')
                    col += 1
                for cat in categories:
                    val = item['income_production_area'][cat]
                    ws.cell(row=current_row, column=col, value=round(val, 2) if val != 0 else '')
                    col += 1
                # На склад
                for cat in categories:
                    ws.cell(row=current_row, column=col,
                            value=item['to_stock'][cat] if item['to_stock'][cat] != 0 else '')
                    col += 1
                for cat in categories:
                    val = item['to_stock_volume'][cat]
                    ws.cell(row=current_row, column=col, value=round(val, 3) if val != 0 else '')
                    col += 1
                for cat in categories:
                    val = item['to_stock_area'][cat]
                    ws.cell(row=current_row, column=col, value=round(val, 2) if val != 0 else '')
                    col += 1
                # В магазин
                for cat in categories:
                    ws.cell(row=current_row, column=col,
                            value=item['to_shop'][cat] if item['to_shop'][cat] != 0 else '')
                    col += 1
                for cat in categories:
                    val = item['to_shop_volume'][cat]
                    ws.cell(row=current_row, column=col, value=round(val, 3) if val != 0 else '')
                    col += 1
                for cat in categories:
                    val = item['to_shop_area'][cat]
                    ws.cell(row=current_row, column=col, value=round(val, 2) if val != 0 else '')
                    col += 1
                # Остаток на конец
                for cat in categories:
                    ws.cell(row=current_row, column=col, value=item['ending'][cat] if item['ending'][cat] != 0 else '')
                    col += 1
                for cat in categories:
                    val = item['ending_volume'][cat]
                    ws.cell(row=current_row, column=col, value=round(val, 3) if val != 0 else '')
                    col += 1
                for cat in categories:
                    val = item['ending_area'][cat]
                    ws.cell(row=current_row, column=col, value=round(val, 2) if val != 0 else '')
                    col += 1

                # Применяем границы и выравнивание
                for c in range(1, 65):
                    cell = ws.cell(row=current_row, column=c)
                    cell.border = thin_border
                    if c >= 5:
                        cell.alignment = Alignment(horizontal='right')
                current_row += 1

            # Итог по группе
            grp = group
            ws.cell(row=current_row, column=1, value=f"ИТОГО: {grp['product_name']} ({species})")
            ws.cell(row=current_row, column=1).font = Font(bold=True)

            col = 5
            # Остаток на начало итого
            for cat in categories:
                ws.cell(row=current_row, column=col,
                        value=grp['total_initial'][cat] if grp['total_initial'][cat] != 0 else '')
                col += 1
            for cat in categories:
                val = grp['total_initial_volume'][cat]
                ws.cell(row=current_row, column=col, value=round(val, 3) if val != 0 else '')
                col += 1
            for cat in categories:
                val = grp['total_initial_area'][cat]
                ws.cell(row=current_row, column=col, value=round(val, 2) if val != 0 else '')
                col += 1
            # Поступление итого
            for cat in categories:
                ws.cell(row=current_row, column=col,
                        value=grp['total_income_production'][cat] if grp['total_income_production'][cat] != 0 else '')
                col += 1
            for cat in categories:
                val = grp['total_income_production_volume'][cat]
                ws.cell(row=current_row, column=col, value=round(val, 3) if val != 0 else '')
                col += 1
            for cat in categories:
                val = grp['total_income_production_area'][cat]
                ws.cell(row=current_row, column=col, value=round(val, 2) if val != 0 else '')
                col += 1
            # На склад итого
            for cat in categories:
                ws.cell(row=current_row, column=col,
                        value=grp['total_to_stock'][cat] if grp['total_to_stock'][cat] != 0 else '')
                col += 1
            for cat in categories:
                val = grp['total_to_stock_volume'][cat]
                ws.cell(row=current_row, column=col, value=round(val, 3) if val != 0 else '')
                col += 1
            for cat in categories:
                val = grp['total_to_stock_area'][cat]
                ws.cell(row=current_row, column=col, value=round(val, 2) if val != 0 else '')
                col += 1
            # В магазин итого
            for cat in categories:
                ws.cell(row=current_row, column=col,
                        value=grp['total_to_shop'][cat] if grp['total_to_shop'][cat] != 0 else '')
                col += 1
            for cat in categories:
                val = grp['total_to_shop_volume'][cat]
                ws.cell(row=current_row, column=col, value=round(val, 3) if val != 0 else '')
                col += 1
            for cat in categories:
                val = grp['total_to_shop_area'][cat]
                ws.cell(row=current_row, column=col, value=round(val, 2) if val != 0 else '')
                col += 1
            # Остаток на конец итого
            for cat in categories:
                ws.cell(row=current_row, column=col,
                        value=grp['total_ending'][cat] if grp['total_ending'][cat] != 0 else '')
                col += 1
            for cat in categories:
                val = grp['total_ending_volume'][cat]
                ws.cell(row=current_row, column=col, value=round(val, 3) if val != 0 else '')
                col += 1
            for cat in categories:
                val = grp['total_ending_area'][cat]
                ws.cell(row=current_row, column=col, value=round(val, 2) if val != 0 else '')
                col += 1

            for c in range(1, 65):
                cell = ws.cell(row=current_row, column=c)
                cell.border = thin_border
                cell.font = Font(bold=True)
                cell.fill = total_fill
                if c >= 5:
                    cell.alignment = Alignment(horizontal='right')
            current_row += 1

    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    for col in range(5, 65):
        ws.column_dimensions[get_column_letter(col)].width = 9

    # Формируем ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{title}_{date_from}_{date_to}.xlsx"'
    wb.save(response)
    return response

def report_category_unit(request):
    """Сводный отчет по категориям для штучных изделий"""
    if request.GET.get('date_from') and request.GET.get('date_to'):
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        return redirect(f'/reports/category-unit/result/?date_from={date_from}&date_to={date_to}')

    return render(request, 'lumber_track/report_form.html', {
        'title': 'Сводный отчет по категориям (штучные)',
        'url_name': 'lumber_track:report_category_unit_result'
    })


def report_category_unit_result(request):
    """Результат сводного отчета по категориям для штучных изделий"""
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if not date_from or not date_to:
        return redirect('lumber_track:report_category_unit')

    from django.db.models import Sum, Q
    from datetime import datetime

    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()

    # Получаем все уникальные позиции ТОЛЬКО ДЛЯ ШТУЧНЫХ
    all_items = DocumentItem.objects.select_related(
        'product_name', 'species', 'grade', 'lumber_dim', 'unit_dim'
    ).filter(
        document__doc_date__lte=date_to_obj,
        unit_dim__isnull=False  # Только штучные (есть unit_dim)
    ).exclude(
        unit_dim_id__isnull=True
    )

    # Собираем данные по каждой позиции
    report_data = {}
    categories = ['A', 'B', 'C', 'D']

    for item in all_items:
        # Для штучных - размеры из unit_dim
        if item.unit_dim:
            length = item.unit_dim.length
            width = item.unit_dim.width
            height = item.unit_dim.height
            dimension_display = f"{length}×{width}×{height}"
            dimension_type = 'unit'
            dimension_id = item.unit_dim.id
        else:
            length = width = height = 0
            dimension_display = "—"
            dimension_type = None
            dimension_id = None

        key = f"{item.product_name_id}_{dimension_id}_{dimension_type}"

        if key not in report_data:
            report_data[key] = {
                'product_name': item.product_name.name,
                'species': item.species.name,
                'length': length,
                'width': width,
                'height': height,
                'dimension_display': dimension_display,
                'dimension_type': dimension_type,
                'dimension_id': dimension_id,
                'initial': {cat: 0 for cat in categories},
                'income_production': {cat: 0 for cat in categories},
                'to_stock': {cat: 0 for cat in categories},
                'to_shop': {cat: 0 for cat in categories},
                'ending': {cat: 0 for cat in categories},
            }

    # Собираем данные по каждому ключу
    for key, data in report_data.items():
        product_name = data['product_name']
        dimension_id = data['dimension_id']
        dimension_type = data['dimension_type']

        for grade_code in categories:
            grade = QualityGrade.objects.filter(code=grade_code).first()
            if not grade:
                continue
            grade_id = grade.id

            # Остаток на начало
            initial_filter = Q(document__doc_date__lt=date_from_obj) & (
                    (Q(document__doc_type=1) & Q(document__location_id=2)) |
                    (Q(document__doc_type=2) & Q(document__location_id=2)) |
                    (Q(document__doc_type=3) & Q(document__to_location_id=2))
            ) & Q(product_name__name=product_name, grade_id=grade_id)

            # Приход из цеха
            income_production_filter = Q(
                document__doc_date__gte=date_from_obj,
                document__doc_date__lte=date_to_obj,
                document__doc_type=2,
                document__location_id=2
            ) & Q(product_name__name=product_name, grade_id=grade_id)

            # Передано на склад (расходные, to_location=розничный склад id=4)
            to_stock_filter = Q(
                document__doc_date__gte=date_from_obj,
                document__doc_date__lte=date_to_obj,
                document__doc_type=3,
                document__to_location_id=4
            ) & Q(product_name__name=product_name, grade_id=grade_id)

            # Передано в магазин (расходные, to_location=ТД Ангара id=3)
            to_shop_filter = Q(
                document__doc_date__gte=date_from_obj,
                document__doc_date__lte=date_to_obj,
                document__doc_type=3,
                document__to_location_id=3
            ) & Q(product_name__name=product_name, grade_id=grade_id)

            # Фильтрация по размерам
            if dimension_type == 'unit' and dimension_id:
                initial_filter &= Q(unit_dim_id=dimension_id)
                income_production_filter &= Q(unit_dim_id=dimension_id)
                to_stock_filter &= Q(unit_dim_id=dimension_id)
                to_shop_filter &= Q(unit_dim_id=dimension_id)

            # Получаем суммы
            initial_qty = DocumentItem.objects.filter(initial_filter).aggregate(total=Sum('quantity'))['total'] or 0
            income_qty = DocumentItem.objects.filter(income_production_filter).aggregate(total=Sum('quantity'))[
                             'total'] or 0
            to_stock_qty = DocumentItem.objects.filter(to_stock_filter).aggregate(total=Sum('quantity'))['total'] or 0
            to_shop_qty = DocumentItem.objects.filter(to_shop_filter).aggregate(total=Sum('quantity'))['total'] or 0

            data['initial'][grade_code] = initial_qty
            data['income_production'][grade_code] = income_qty
            data['to_stock'][grade_code] = to_stock_qty
            data['to_shop'][grade_code] = to_shop_qty
            data['ending'][grade_code] = initial_qty + income_qty - to_stock_qty - to_shop_qty

    # Фильтруем позиции с данными
    report_list = []
    for key, data in report_data.items():
        has_data = False
        for cat in categories:
            if (data['initial'][cat] != 0 or
                    data['income_production'][cat] != 0 or
                    data['to_stock'][cat] != 0 or
                    data['to_shop'][cat] != 0 or
                    data['ending'][cat] != 0):
                has_data = True
                break
        if has_data:
            report_list.append(data)

    if not report_list:
        context = {
            'title': 'Сводный отчет по категориям (штучные)',
            'date_from': date_from_obj,
            'date_to': date_to_obj,
            'grouped_by_species': {},
            'categories': categories,
            'url_name': 'lumber_track:report_category_unit_result',
        }
        return render(request, 'lumber_track/report_category.html', context)

    report_list.sort(key=lambda x: x['product_name'])

    # Группировка по породе и наименованию
    grouped_by_species = {}
    for item in report_list:
        species = item['species']
        product_name = item['product_name']

        if species not in grouped_by_species:
            grouped_by_species[species] = {}

        if product_name not in grouped_by_species[species]:
            grouped_by_species[species][product_name] = {
                'product_name': product_name,
                'species': species,
                'items': [],
                'total_initial': {cat: 0 for cat in categories},
                'total_income_production': {cat: 0 for cat in categories},
                'total_to_stock': {cat: 0 for cat in categories},
                'total_to_shop': {cat: 0 for cat in categories},
                'total_ending': {cat: 0 for cat in categories},
            }

        grouped_by_species[species][product_name]['items'].append(item)

        for cat in categories:
            grouped_by_species[species][product_name]['total_initial'][cat] += item['initial'][cat]
            grouped_by_species[species][product_name]['total_income_production'][cat] += item['income_production'][cat]
            grouped_by_species[species][product_name]['total_to_stock'][cat] += item['to_stock'][cat]
            grouped_by_species[species][product_name]['total_to_shop'][cat] += item['to_shop'][cat]
            grouped_by_species[species][product_name]['total_ending'][cat] += item['ending'][cat]

    context = {
        'title': 'Сводный отчет по категориям (штучные)',
        'date_from': date_from_obj,
        'date_to': date_to_obj,
        'grouped_by_species': grouped_by_species,
        'categories': categories,
        'url_name': 'lumber_track:report_category_unit_result',
    }

    if request.GET.get('export') == 'excel':
        return export_category_unit_to_excel(grouped_by_species, categories, context['title'], date_from, date_to)

    return render(request, 'lumber_track/report_category.html', context)


def export_category_unit_to_excel(grouped_by_species, categories, title, date_from, date_to):
    """Экспорт сводного отчета для штучных изделий в Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    subheader_font = Font(bold=True, size=10)
    subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    species_fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')
    total_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 1. Первая строка заголовков (основные секции)
    col = 1
    ws.cell(row=1, column=col, value='Наименование')
    ws.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
    col += 1

    ws.cell(row=1, column=col, value='Размер, мм')
    ws.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col + 2)
    col += 3

    sections = ['Остаток на начало', 'Поступление из цеха', 'Передано на склад', 'Передано в магазин',
                'Остаток на конец']
    for section in sections:
        ws.cell(row=1, column=col, value=section)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
        col += 4

    # 2. Вторая строка заголовков (показатели)
    col = 5
    for _ in range(5):
        ws.cell(row=2, column=col, value='Кол-во (шт)')
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 3)
        col += 4

    # 3. Третья строка заголовков (категории)
    col = 5
    for _ in range(5):
        for cat in categories:
            ws.cell(row=3, column=col, value=cat)
            col += 1

    # Оформление заголовков
    for row in range(1, 4):
        for col in range(1, 4 + 5 * 4):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if row == 1:
                cell.font = header_font
                cell.fill = header_fill
            elif row == 2:
                cell.font = subheader_font
                cell.fill = subheader_fill
            else:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')

    # 4. Данные
    current_row = 4
    for species, groups in grouped_by_species.items():
        # Заголовок породы
        cell = ws.cell(row=current_row, column=1, value=f"🌲 {species}")
        cell.font = Font(bold=True, size=11)
        cell.fill = species_fill
        for col in range(1, 4 + 5 * 4):
            ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1

        for group in groups.values():
            # Детальные строки
            for item in group['items']:
                ws.cell(row=current_row, column=1, value=item['product_name'])
                ws.cell(row=current_row, column=2, value=item['length'] if item['length'] else '-')
                ws.cell(row=current_row, column=3, value=item['width'] if item['width'] else '-')
                ws.cell(row=current_row, column=4, value=item['height'] if item['height'] else '-')

                col = 5
                # Остаток на начало
                for cat in categories:
                    ws.cell(row=current_row, column=col,
                            value=item['initial'][cat] if item['initial'][cat] != 0 else '')
                    col += 1
                # Поступление
                for cat in categories:
                    ws.cell(row=current_row, column=col,
                            value=item['income_production'][cat] if item['income_production'][cat] != 0 else '')
                    col += 1
                # На склад
                for cat in categories:
                    ws.cell(row=current_row, column=col,
                            value=item['to_stock'][cat] if item['to_stock'][cat] != 0 else '')
                    col += 1
                # В магазин
                for cat in categories:
                    ws.cell(row=current_row, column=col,
                            value=item['to_shop'][cat] if item['to_shop'][cat] != 0 else '')
                    col += 1
                # Остаток на конец
                for cat in categories:
                    ws.cell(row=current_row, column=col, value=item['ending'][cat] if item['ending'][cat] != 0 else '')
                    col += 1

                # Границы и выравнивание
                for c in range(1, 4 + 5 * 4):
                    cell = ws.cell(row=current_row, column=c)
                    cell.border = thin_border
                    if c >= 5:
                        cell.alignment = Alignment(horizontal='right')
                current_row += 1

            # Итог по группе
            grp = group
            ws.cell(row=current_row, column=1, value=f"ИТОГО: {grp['product_name']} ({species})")
            ws.cell(row=current_row, column=1).font = Font(bold=True)

            col = 5
            # Остаток на начало итого
            for cat in categories:
                ws.cell(row=current_row, column=col,
                        value=grp['total_initial'][cat] if grp['total_initial'][cat] != 0 else '')
                col += 1
            # Поступление итого
            for cat in categories:
                ws.cell(row=current_row, column=col,
                        value=grp['total_income_production'][cat] if grp['total_income_production'][cat] != 0 else '')
                col += 1
            # На склад итого
            for cat in categories:
                ws.cell(row=current_row, column=col,
                        value=grp['total_to_stock'][cat] if grp['total_to_stock'][cat] != 0 else '')
                col += 1
            # В магазин итого
            for cat in categories:
                ws.cell(row=current_row, column=col,
                        value=grp['total_to_shop'][cat] if grp['total_to_shop'][cat] != 0 else '')
                col += 1
            # Остаток на конец итого
            for cat in categories:
                ws.cell(row=current_row, column=col,
                        value=grp['total_ending'][cat] if grp['total_ending'][cat] != 0 else '')
                col += 1

            for c in range(1, 4 + 5 * 4):
                cell = ws.cell(row=current_row, column=c)
                cell.border = thin_border
                cell.font = Font(bold=True)
                cell.fill = total_fill
                if c >= 5:
                    cell.alignment = Alignment(horizontal='right')
            current_row += 1

    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    for col in range(5, 4 + 5 * 4):
        ws.column_dimensions[get_column_letter(col)].width = 12

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{title}_{date_from}_{date_to}.xlsx"'
    wb.save(response)
    return response

def report_stock_balance(request):
    """Отчет по остаткам на складе (без выбора периода)"""
    # Сразу перенаправляем на результат
    return redirect('lumber_track:report_stock_balance_result')


def report_stock_balance_result(request):
    """Результат отчета по остаткам на складе (на текущую дату)"""
    from collections import defaultdict
    from django.utils import timezone
    from django.db.models import Sum

    date_to_obj = timezone.now().date()

    # Получаем все позиции из документов
    # Создаем словарь для агрегации
    stocks = {}

    # 1. Начальные остатки + Приходы (увеличивают остаток)
    incoming = DocumentItem.objects.filter(
        document__doc_type__in=[1, 2],
        document__location_id=2,
        document__doc_date__lte=date_to_obj
    ).select_related('product_name', 'species', 'grade', 'lumber_dim', 'unit_dim')

    for item in incoming:
        key = f"{item.product_name_id}_{item.species_id}_{item.grade_id}_{item.lumber_dim_id}_{item.unit_dim_id}"

        if key not in stocks:
            if item.lumber_dim:
                dim_display = f"{item.lumber_dim.thickness}×{item.lumber_dim.width}×{item.lumber_dim.length}"
                volume_per_unit = item.lumber_dim.volume_m3
                area_per_unit = item.lumber_dim.area_m2
            elif item.unit_dim:
                dim_display = f"{item.unit_dim.length}×{item.unit_dim.width}×{item.unit_dim.height}"
                volume_per_unit = 0
                area_per_unit = 0
            else:
                dim_display = "—"
                volume_per_unit = 0
                area_per_unit = 0

            stocks[key] = {
                'product_name': item.product_name.name,
                'species': item.species.name,
                'grade': item.grade.code,
                'dimension_display': dim_display,
                'quantity': 0,
                'volume': 0,
                'area': 0,
                'volume_per_unit': volume_per_unit,
                'area_per_unit': area_per_unit,
            }

        stocks[key]['quantity'] += item.quantity

    # 2. Расходы (уменьшают остаток)
    outgoing = DocumentItem.objects.filter(
        document__doc_type=3,
        document__doc_date__lte=date_to_obj
    ).exclude(
        document__to_location_id=2
    ).select_related('product_name', 'species', 'grade', 'lumber_dim', 'unit_dim')

    for item in outgoing:
        key = f"{item.product_name_id}_{item.species_id}_{item.grade_id}_{item.lumber_dim_id}_{item.unit_dim_id}"

        if key not in stocks:
            if item.lumber_dim:
                dim_display = f"{item.lumber_dim.thickness}×{item.lumber_dim.width}×{item.lumber_dim.length}"
                volume_per_unit = item.lumber_dim.volume_m3
                area_per_unit = item.lumber_dim.area_m2
            elif item.unit_dim:
                dim_display = f"{item.unit_dim.length}×{item.unit_dim.width}×{item.unit_dim.height}"
                volume_per_unit = 0
                area_per_unit = 0
            else:
                dim_display = "—"
                volume_per_unit = 0
                area_per_unit = 0

            stocks[key] = {
                'product_name': item.product_name.name,
                'species': item.species.name,
                'grade': item.grade.code,
                'dimension_display': dim_display,
                'quantity': 0,
                'volume': 0,
                'area': 0,
                'volume_per_unit': volume_per_unit,
                'area_per_unit': area_per_unit,
            }

        stocks[key]['quantity'] -= item.quantity

    # Формируем итоговый список
    report_list = []
    for key, data in stocks.items():
        if data['quantity'] > 0:
            data['volume'] = data['quantity'] * data['volume_per_unit']
            data['area'] = data['quantity'] * data['area_per_unit']
            report_list.append(data)

    if not report_list:
        context = {
            'title': 'Остатки на складе готовой продукции',
            'report_date': date_to_obj,
            'grouped_data': {},
            'grand_total_quantity': 0,
            'grand_total_volume': 0,
            'grand_total_area': 0,
            'url_name': 'lumber_track:report_stock_balance_result',
        }
        return render(request, 'lumber_track/report_stock_balance.html', context)

    # Группировка по породе и категории
    grouped_data = {}
    for item in report_list:
        species = item['species']
        grade = item['grade']

        if species not in grouped_data:
            grouped_data[species] = {}

        if grade not in grouped_data[species]:
            grouped_data[species][grade] = []

        grouped_data[species][grade].append(item)

    # Общие итоги
    grand_total_quantity = sum(item['quantity'] for item in report_list)
    grand_total_volume = sum(item['volume'] for item in report_list)
    grand_total_area = sum(item['area'] for item in report_list)

    context = {
        'title': 'Остатки на складе готовой продукции',
        'report_date': date_to_obj,
        'grouped_data': grouped_data,
        'grand_total_quantity': grand_total_quantity,
        'grand_total_volume': grand_total_volume,
        'grand_total_area': grand_total_area,
        'url_name': 'lumber_track:report_stock_balance_result',
    }

    if request.GET.get('export') == 'excel':
        return export_stock_balance_to_excel(grouped_data, context['title'], date_to_obj,
                                             grand_total_quantity, grand_total_volume, grand_total_area)

    return render(request, 'lumber_track/report_stock_balance.html', context)

def get_item_key(item):
    """Формирует уникальный ключ для позиции"""
    return f"{item.product_name_id}_{item.species_id}_{item.grade_id}_{item.lumber_dim_id}_{item.unit_dim_id}"


def create_item_dict(item):
    """Создает словарь с данными позиции"""
    if item.lumber_dim:
        dimension_display = f"{item.lumber_dim.thickness}×{item.lumber_dim.width}×{item.lumber_dim.length}"
    elif item.unit_dim:
        dimension_display = f"{item.unit_dim.length}×{item.unit_dim.width}×{item.unit_dim.height}"
    else:
        dimension_display = "—"

    return {
        'product_name': item.product_name.name,
        'species': item.species.name,
        'grade': item.grade.code,
        'dimension_display': dimension_display,
        'quantity': 0,
        'volume': 0,
        'area': 0,
    }

def report_detailed(request):
    """Детальный отчет по документам (разбивка по дням)"""
    if request.GET.get('date_from') and request.GET.get('date_to'):
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        return redirect(f'/reports/detailed/result/?date_from={date_from}&date_to={date_to}')

    return render(request, 'lumber_track/report_form.html', {
        'title': 'Детальный отчет по дням',
        'url_name': 'lumber_track:report_detailed_result'
    })


def report_detailed_result(request):
    """Результат детального отчета по дням"""
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if not date_from or not date_to:
        return redirect('lumber_track:report_detailed')

    from datetime import datetime

    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()

    # Получаем все позиции из документов за период с сортировкой по дате
    items = DocumentItem.objects.filter(
        document__doc_date__gte=date_from_obj,
        document__doc_date__lte=date_to_obj
    ).select_related(
        'document', 'product_name', 'species', 'grade', 'lumber_dim', 'unit_dim'
    ).order_by('document__doc_date', 'document__doc_number')

    report_data = []
    for item in items:
        doc = item.document
        doc_type_name = dict(Document.DOCUMENT_TYPES).get(doc.doc_type, '')

        # Получаем размер
        if item.lumber_dim:
            dimension_display = f"{item.lumber_dim.thickness}-{item.lumber_dim.width}-{item.lumber_dim.length} мм"
        elif item.unit_dim:
            dimension_display = f"{item.unit_dim.length}-{item.unit_dim.width}-{item.unit_dim.height} мм"
        else:
            dimension_display = "—"

        report_data.append({
            'doc_date': doc.doc_date,
            'doc_number': doc.doc_number,
            'doc_type': doc_type_name,
            'product_name': item.product_name.name,
            'species': item.species.name,
            'grade': item.grade.code,
            'dimension_display': dimension_display,
            'quantity': item.quantity,
            'volume': item.volume_m3,
            'area': item.area_m2,
        })

    # Подсчет итогов
    total_quantity = sum(item['quantity'] for item in report_data)
    total_volume = sum(item['volume'] for item in report_data)
    total_area = sum(item['area'] for item in report_data)

    context = {
        'title': 'Детальный отчет по дням',
        'date_from': date_from_obj,
        'date_to': date_to_obj,
        'report_data': report_data,
        'total_quantity': total_quantity,
        'total_volume': total_volume,
        'total_area': total_area,
        'url_name': 'lumber_track:report_detailed_result',
    }

    if request.GET.get('export') == 'excel':
        return export_detailed_to_excel(report_data, context['title'], date_from, date_to,
                                        total_quantity, total_volume, total_area)

    return render(request, 'lumber_track/report_detailed.html', context)


def export_detailed_to_excel(report_data, title, date_from, date_to,
                             total_quantity, total_volume, total_area):
    """Экспорт детального отчета в Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Детальный отчет"

    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    # Заголовки
    headers = ['Дата', '№ документа', 'Тип документа', 'Наименование', 'Порода',
               'Категория', 'Размер', 'Количество (шт)', 'Объем (м³)', 'Площадь (м²)']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.fill = header_fill

    # Данные
    for row, item in enumerate(report_data, 2):
        ws.cell(row=row, column=1, value=item['doc_date'].strftime('%d.%m.%Y'))
        ws.cell(row=row, column=2, value=item['doc_number'])
        ws.cell(row=row, column=3, value=item['doc_type'])
        ws.cell(row=row, column=4, value=item['product_name'])
        ws.cell(row=row, column=5, value=item['species'])
        ws.cell(row=row, column=6, value=item['grade'])
        ws.cell(row=row, column=7, value=item['dimension_display'])
        ws.cell(row=row, column=8, value=float(item['quantity']))
        ws.cell(row=row, column=9, value=float(item['volume']))
        ws.cell(row=row, column=10, value=float(item['area']))

    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{title}_{date_from}_{date_to}.xlsx"'
    wb.save(response)
    return response


def export_income_to_excel(report_data, title, date_from, date_to, grand_total_quantity, grand_total_volume,
                           grand_total_area):
    """Экспорт отчета в Excel с группировкой по дням"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel ограничение на длину имени листа

    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    date_header_fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')
    date_header_font = Font(bold=True, size=11)
    day_total_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    day_total_font = Font(bold=True)
    total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    total_font = Font(bold=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовки
    headers = ['Наименование', 'Порода', 'Категория', 'Размер', 'Количество (шт)', 'Объем (м³)', 'Площадь (м²)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = thin_border

    current_row = 2

    for item in report_data:
        if item.get('is_date_header'):
            # Заголовок даты
            cell = ws.cell(row=current_row, column=1, value=f"📅 {item['date'].strftime('%d.%m.%Y')}")
            cell.font = date_header_font
            cell.fill = date_header_fill
            cell.border = thin_border
            # Объединяем ячейки для заголовка даты
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            current_row += 1

        elif item.get('is_day_total'):
            # Итоги за день
            ws.cell(row=current_row, column=1, value="Итого за день:")
            ws.cell(row=current_row, column=1).font = day_total_font
            ws.cell(row=current_row, column=1).fill = day_total_fill
            ws.cell(row=current_row, column=1).border = thin_border
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='right')

            ws.cell(row=current_row, column=5, value=float(item['total_quantity']))
            ws.cell(row=current_row, column=5).font = day_total_font
            ws.cell(row=current_row, column=5).fill = day_total_fill
            ws.cell(row=current_row, column=5).border = thin_border
            ws.cell(row=current_row, column=5).alignment = Alignment(horizontal='right')

            ws.cell(row=current_row, column=6, value=float(item['total_volume']))
            ws.cell(row=current_row, column=6).font = day_total_font
            ws.cell(row=current_row, column=6).fill = day_total_fill
            ws.cell(row=current_row, column=6).border = thin_border
            ws.cell(row=current_row, column=6).alignment = Alignment(horizontal='right')

            ws.cell(row=current_row, column=7, value=float(item['total_area']))
            ws.cell(row=current_row, column=7).font = day_total_font
            ws.cell(row=current_row, column=7).fill = day_total_fill
            ws.cell(row=current_row, column=7).border = thin_border
            ws.cell(row=current_row, column=7).alignment = Alignment(horizontal='right')

            # Объединяем ячейки для первых 4 колонок
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            current_row += 1

        else:
            # Детальные строки
            ws.cell(row=current_row, column=1, value=item['product_name'])
            ws.cell(row=current_row, column=2, value=item['species'])
            ws.cell(row=current_row, column=3, value=item['grade'])
            ws.cell(row=current_row, column=4, value=item['dimension_display'])
            ws.cell(row=current_row, column=5, value=float(item['total_quantity']))
            ws.cell(row=current_row, column=6, value=float(item['total_volume']))
            ws.cell(row=current_row, column=7, value=float(item['total_area']))

            for col in range(1, 8):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                if col >= 5:
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')
            current_row += 1

    # Общие итоги
    ws.cell(row=current_row, column=1, value="ВСЕГО за период:")
    ws.cell(row=current_row, column=1).font = total_font
    ws.cell(row=current_row, column=1).fill = total_fill
    ws.cell(row=current_row, column=1).border = thin_border
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='right')

    ws.cell(row=current_row, column=5, value=float(grand_total_quantity))
    ws.cell(row=current_row, column=5).font = total_font
    ws.cell(row=current_row, column=5).fill = total_fill
    ws.cell(row=current_row, column=5).border = thin_border
    ws.cell(row=current_row, column=5).alignment = Alignment(horizontal='right')

    ws.cell(row=current_row, column=6, value=float(grand_total_volume))
    ws.cell(row=current_row, column=6).font = total_font
    ws.cell(row=current_row, column=6).fill = total_fill
    ws.cell(row=current_row, column=6).border = thin_border
    ws.cell(row=current_row, column=6).alignment = Alignment(horizontal='right')

    ws.cell(row=current_row, column=7, value=float(grand_total_area))
    ws.cell(row=current_row, column=7).font = total_font
    ws.cell(row=current_row, column=7).fill = total_fill
    ws.cell(row=current_row, column=7).border = thin_border
    ws.cell(row=current_row, column=7).alignment = Alignment(horizontal='right')

    # Объединяем ячейки для итогов (первые 4 колонки)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)

    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    # Формируем ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{title}_{date_from}_{date_to}.xlsx"'
    wb.save(response)
    return response


def export_to_shop_excel(report_data, title, date_from, date_to, grand_total_quantity, grand_total_volume,
                         grand_total_area):
    """Экспорт отчета 'В магазин' в Excel с группировкой по дням"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Стили (аналогично экспорту для поступления)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    date_header_fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')
    date_header_font = Font(bold=True, size=11)
    day_total_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    day_total_font = Font(bold=True)
    total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    total_font = Font(bold=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовки
    headers = ['Наименование', 'Порода', 'Категория', 'Размер', 'Количество (шт)', 'Объем (м³)', 'Площадь (м²)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = thin_border

    current_row = 2

    for item in report_data:
        if item.get('is_date_header'):
            cell = ws.cell(row=current_row, column=1, value=f"📅 {item['date'].strftime('%d.%m.%Y')}")
            cell.font = date_header_font
            cell.fill = date_header_fill
            cell.border = thin_border
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            current_row += 1

        elif item.get('is_day_total'):
            ws.cell(row=current_row, column=1, value="Итого за день:")
            ws.cell(row=current_row, column=1).font = day_total_font
            ws.cell(row=current_row, column=1).fill = day_total_fill
            ws.cell(row=current_row, column=1).border = thin_border
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='right')

            ws.cell(row=current_row, column=5, value=float(item['total_quantity']))
            ws.cell(row=current_row, column=5).font = day_total_font
            ws.cell(row=current_row, column=5).fill = day_total_fill
            ws.cell(row=current_row, column=5).border = thin_border
            ws.cell(row=current_row, column=5).alignment = Alignment(horizontal='right')

            ws.cell(row=current_row, column=6, value=float(item['total_volume']))
            ws.cell(row=current_row, column=6).font = day_total_font
            ws.cell(row=current_row, column=6).fill = day_total_fill
            ws.cell(row=current_row, column=6).border = thin_border
            ws.cell(row=current_row, column=6).alignment = Alignment(horizontal='right')

            ws.cell(row=current_row, column=7, value=float(item['total_area']))
            ws.cell(row=current_row, column=7).font = day_total_font
            ws.cell(row=current_row, column=7).fill = day_total_fill
            ws.cell(row=current_row, column=7).border = thin_border
            ws.cell(row=current_row, column=7).alignment = Alignment(horizontal='right')

            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            current_row += 1

        else:
            ws.cell(row=current_row, column=1, value=item['product_name'])
            ws.cell(row=current_row, column=2, value=item['species'])
            ws.cell(row=current_row, column=3, value=item['grade'])
            ws.cell(row=current_row, column=4, value=item['dimension_display'])
            ws.cell(row=current_row, column=5, value=float(item['total_quantity']))
            ws.cell(row=current_row, column=6, value=float(item['total_volume']))
            ws.cell(row=current_row, column=7, value=float(item['total_area']))

            for col in range(1, 8):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                if col >= 5:
                    cell.alignment = Alignment(horizontal='right')
            current_row += 1

    # Общие итоги
    ws.cell(row=current_row, column=1, value="ВСЕГО за период:")
    ws.cell(row=current_row, column=1).font = total_font
    ws.cell(row=current_row, column=1).fill = total_fill
    ws.cell(row=current_row, column=1).border = thin_border
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='right')

    ws.cell(row=current_row, column=5, value=float(grand_total_quantity))
    ws.cell(row=current_row, column=5).font = total_font
    ws.cell(row=current_row, column=5).fill = total_fill
    ws.cell(row=current_row, column=5).border = thin_border
    ws.cell(row=current_row, column=5).alignment = Alignment(horizontal='right')

    ws.cell(row=current_row, column=6, value=float(grand_total_volume))
    ws.cell(row=current_row, column=6).font = total_font
    ws.cell(row=current_row, column=6).fill = total_fill
    ws.cell(row=current_row, column=6).border = thin_border
    ws.cell(row=current_row, column=6).alignment = Alignment(horizontal='right')

    ws.cell(row=current_row, column=7, value=float(grand_total_area))
    ws.cell(row=current_row, column=7).font = total_font
    ws.cell(row=current_row, column=7).fill = total_fill
    ws.cell(row=current_row, column=7).border = thin_border
    ws.cell(row=current_row, column=7).alignment = Alignment(horizontal='right')

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)

    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{title}_{date_from}_{date_to}.xlsx"'
    wb.save(response)
    return response


def export_to_stock_excel(report_data, title, date_from, date_to, grand_total_quantity, grand_total_volume,
                          grand_total_area):
    """Экспорт отчета 'На склад' в Excel с группировкой по дням"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    date_header_fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')
    date_header_font = Font(bold=True, size=11)
    day_total_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    day_total_font = Font(bold=True)
    total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    total_font = Font(bold=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовки
    headers = ['Наименование', 'Порода', 'Категория', 'Размер', 'Количество (шт)', 'Объем (м³)', 'Площадь (м²)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = thin_border

    current_row = 2

    for item in report_data:
        if item.get('is_date_header'):
            cell = ws.cell(row=current_row, column=1, value=f"📅 {item['date'].strftime('%d.%m.%Y')}")
            cell.font = date_header_font
            cell.fill = date_header_fill
            cell.border = thin_border
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            current_row += 1

        elif item.get('is_day_total'):
            ws.cell(row=current_row, column=1, value="Итого за день:")
            ws.cell(row=current_row, column=1).font = day_total_font
            ws.cell(row=current_row, column=1).fill = day_total_fill
            ws.cell(row=current_row, column=1).border = thin_border
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='right')

            ws.cell(row=current_row, column=5, value=float(item['total_quantity']))
            ws.cell(row=current_row, column=5).font = day_total_font
            ws.cell(row=current_row, column=5).fill = day_total_fill
            ws.cell(row=current_row, column=5).border = thin_border
            ws.cell(row=current_row, column=5).alignment = Alignment(horizontal='right')

            ws.cell(row=current_row, column=6, value=float(item['total_volume']))
            ws.cell(row=current_row, column=6).font = day_total_font
            ws.cell(row=current_row, column=6).fill = day_total_fill
            ws.cell(row=current_row, column=6).border = thin_border
            ws.cell(row=current_row, column=6).alignment = Alignment(horizontal='right')

            ws.cell(row=current_row, column=7, value=float(item['total_area']))
            ws.cell(row=current_row, column=7).font = day_total_font
            ws.cell(row=current_row, column=7).fill = day_total_fill
            ws.cell(row=current_row, column=7).border = thin_border
            ws.cell(row=current_row, column=7).alignment = Alignment(horizontal='right')

            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            current_row += 1

        else:
            ws.cell(row=current_row, column=1, value=item['product_name'])
            ws.cell(row=current_row, column=2, value=item['species'])
            ws.cell(row=current_row, column=3, value=item['grade'])
            ws.cell(row=current_row, column=4, value=item['dimension_display'])
            ws.cell(row=current_row, column=5, value=float(item['total_quantity']))
            ws.cell(row=current_row, column=6, value=float(item['total_volume']))
            ws.cell(row=current_row, column=7, value=float(item['total_area']))

            for col in range(1, 8):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                if col >= 5:
                    cell.alignment = Alignment(horizontal='right')
            current_row += 1

    # Общие итоги
    ws.cell(row=current_row, column=1, value="ВСЕГО за период:")
    ws.cell(row=current_row, column=1).font = total_font
    ws.cell(row=current_row, column=1).fill = total_fill
    ws.cell(row=current_row, column=1).border = thin_border
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='right')

    ws.cell(row=current_row, column=5, value=float(grand_total_quantity))
    ws.cell(row=current_row, column=5).font = total_font
    ws.cell(row=current_row, column=5).fill = total_fill
    ws.cell(row=current_row, column=5).border = thin_border
    ws.cell(row=current_row, column=5).alignment = Alignment(horizontal='right')

    ws.cell(row=current_row, column=6, value=float(grand_total_volume))
    ws.cell(row=current_row, column=6).font = total_font
    ws.cell(row=current_row, column=6).fill = total_fill
    ws.cell(row=current_row, column=6).border = thin_border
    ws.cell(row=current_row, column=6).alignment = Alignment(horizontal='right')

    ws.cell(row=current_row, column=7, value=float(grand_total_area))
    ws.cell(row=current_row, column=7).font = total_font
    ws.cell(row=current_row, column=7).fill = total_fill
    ws.cell(row=current_row, column=7).border = thin_border
    ws.cell(row=current_row, column=7).alignment = Alignment(horizontal='right')

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)

    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{title}_{date_from}_{date_to}.xlsx"'
    wb.save(response)
    return response


def export_stock_balance_to_excel(grouped_data, title, report_date,
                                  grand_total_quantity, grand_total_volume, grand_total_area):
    """Экспорт отчета по остаткам в Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    species_fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')
    grade_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовки таблицы
    headers = ['Наименование', 'Размер', 'Порода', 'Категория', 'Количество (шт)', 'Объем (м³)', 'Площадь (м²)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = thin_border

    current_row = 2

    # Заполняем данные
    for species, grades in grouped_data.items():
        # Заголовок породы
        cell = ws.cell(row=current_row, column=1, value=f"🌲 {species}")
        cell.font = Font(bold=True, size=11)
        cell.fill = species_fill
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        for col in range(1, 8):
            ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1

        for grade, items in grades.items():
            # Заголовок категории
            cell = ws.cell(row=current_row, column=1, value=f"⭐ Категория {grade}")
            cell.font = Font(bold=True, size=10)
            cell.fill = grade_fill
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1

            for item in items:
                ws.cell(row=current_row, column=1, value=item['product_name'])
                ws.cell(row=current_row, column=2, value=item['dimension_display'])
                ws.cell(row=current_row, column=3, value=item['species'])
                ws.cell(row=current_row, column=4, value=item['grade'])
                ws.cell(row=current_row, column=5, value=float(item['quantity']))
                ws.cell(row=current_row, column=6, value=float(item['volume']))
                ws.cell(row=current_row, column=7, value=float(item['area']))

                for col in range(1, 8):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = thin_border
                    if col >= 5:
                        cell.alignment = Alignment(horizontal='right')
                current_row += 1

    # Общие итоги
    ws.cell(row=current_row, column=1, value="ВСЕГО на складе:")
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=5, value=float(grand_total_quantity))
    ws.cell(row=current_row, column=6, value=float(grand_total_volume))
    ws.cell(row=current_row, column=7, value=float(grand_total_area))
    for col in range(1, 8):
        ws.cell(row=current_row, column=col).border = thin_border
        if col >= 5:
            ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='right')

    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    # Формируем ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{title}_{report_date}.xlsx"'
    wb.save(response)
    return response


# ========== СВОДНЫЕ ОТЧЕТЫ (группировка по породам и категориям) ==========

def report_income_summary(request):
    """Сводный отчет по поступлению (форма выбора периода)"""
    if request.GET.get('date_from') and request.GET.get('date_to'):
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        return redirect(f'/reports/income-summary/result/?date_from={date_from}&date_to={date_to}')

    return render(request, 'lumber_track/report_form.html', {
        'title': 'Поступление от ЦСИ (сводный)',
        'url_name': 'lumber_track:report_income_summary_result'
    })


def report_income_summary_result(request):
    """Результат сводного отчета по поступлению"""
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if not date_from or not date_to:
        return redirect('lumber_track:report_income_summary')

    from datetime import datetime
    from collections import defaultdict

    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()

    # Получаем данные из приходных документов
    items = DocumentItem.objects.filter(
        document__doc_type=2,
        document__location_id=2,
        document__doc_date__gte=date_from_obj,
        document__doc_date__lte=date_to_obj
    ).select_related('product_name', 'species', 'grade', 'lumber_dim', 'unit_dim')

    # Структура: порода -> категория -> список товаров
    report_data = defaultdict(lambda: defaultdict(list))

    for item in items:
        species = item.species.name
        grade = item.grade.code
        product_name = item.product_name.name

        if item.lumber_dim:
            dimension_display = f"{item.lumber_dim.thickness}×{item.lumber_dim.width}×{item.lumber_dim.length}"
        else:
            dimension_display = "—"

        # Ищем существующую запись
        items_list = report_data[species][grade]
        found = False
        for existing in items_list:
            if existing['product_name'] == product_name and existing['dimension_display'] == dimension_display:
                existing['quantity'] += item.quantity
                existing['volume'] += item.volume_m3
                existing['area'] += item.area_m2
                found = True
                break

        if not found:
            items_list.append({
                'product_name': product_name,
                'dimension_display': dimension_display,
                'quantity': item.quantity,
                'volume': item.volume_m3,
                'area': item.area_m2,
            })

    # Подсчет итогов
    summary_data = {}
    grand_total_quantity = 0
    grand_total_volume = 0
    grand_total_area = 0

    for species, grades in report_data.items():
        summary_data[species] = {}
        for grade, items_list in grades.items():
            grade_total_quantity = sum(item['quantity'] for item in items_list)
            grade_total_volume = sum(item['volume'] for item in items_list)
            grade_total_area = sum(item['area'] for item in items_list)

            summary_data[species][grade] = {
                'items': items_list,
                'total_quantity': grade_total_quantity,
                'total_volume': grade_total_volume,
                'total_area': grade_total_area,
            }

            grand_total_quantity += grade_total_quantity
            grand_total_volume += grade_total_volume
            grand_total_area += grade_total_area

    context = {
        'title': 'Поступление от ЦСИ (сводный)',
        'date_from': date_from_obj,
        'date_to': date_to_obj,
        'summary_data': summary_data,
        'grand_total_quantity': grand_total_quantity,
        'grand_total_volume': grand_total_volume,
        'grand_total_area': grand_total_area,
        'url_name': 'lumber_track:report_income_summary_result',
    }

    if request.GET.get('export') == 'excel':
        return export_summary_to_excel(summary_data, context['title'], date_from, date_to,
                                       grand_total_quantity, grand_total_volume, grand_total_area)

    return render(request, 'lumber_track/report_summary.html', context)


def report_to_stock_summary(request):
    """Сводный отчет по отгрузкам на склад (форма выбора периода)"""
    if request.GET.get('date_from') and request.GET.get('date_to'):
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        return redirect(f'/reports/to-stock-summary/result/?date_from={date_from}&date_to={date_to}')

    return render(request, 'lumber_track/report_form.html', {
        'title': 'На склад (сводный)',
        'url_name': 'lumber_track:report_to_stock_summary_result'
    })


def report_to_stock_summary_result(request):
    """Результат сводного отчета по отгрузкам на склад"""
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if not date_from or not date_to:
        return redirect('lumber_track:report_to_stock_summary')

    from datetime import datetime
    from collections import defaultdict

    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()

    items = DocumentItem.objects.filter(
        document__doc_type=3,
        document__to_location_id=4,
        document__doc_date__gte=date_from_obj,
        document__doc_date__lte=date_to_obj
    ).select_related('product_name', 'species', 'grade', 'lumber_dim', 'unit_dim')

    report_data = defaultdict(lambda: defaultdict(list))

    for item in items:
        species = item.species.name
        grade = item.grade.code
        product_name = item.product_name.name

        if item.lumber_dim:
            dimension_display = f"{item.lumber_dim.thickness}×{item.lumber_dim.width}×{item.lumber_dim.length}"
        else:
            dimension_display = "—"

        items_list = report_data[species][grade]
        found = False
        for existing in items_list:
            if existing['product_name'] == product_name and existing['dimension_display'] == dimension_display:
                existing['quantity'] += item.quantity
                existing['volume'] += item.volume_m3
                existing['area'] += item.area_m2
                found = True
                break

        if not found:
            items_list.append({
                'product_name': product_name,
                'dimension_display': dimension_display,
                'quantity': item.quantity,
                'volume': item.volume_m3,
                'area': item.area_m2,
            })

    summary_data = {}
    grand_total_quantity = 0
    grand_total_volume = 0
    grand_total_area = 0

    for species, grades in report_data.items():
        summary_data[species] = {}
        for grade, items_list in grades.items():
            grade_total_quantity = sum(item['quantity'] for item in items_list)
            grade_total_volume = sum(item['volume'] for item in items_list)
            grade_total_area = sum(item['area'] for item in items_list)

            summary_data[species][grade] = {
                'items': items_list,
                'total_quantity': grade_total_quantity,
                'total_volume': grade_total_volume,
                'total_area': grade_total_area,
            }

            grand_total_quantity += grade_total_quantity
            grand_total_volume += grade_total_volume
            grand_total_area += grade_total_area

    context = {
        'title': 'Магазин в Стрелке (сводный)',
        'date_from': date_from_obj,
        'date_to': date_to_obj,
        'summary_data': summary_data,
        'grand_total_quantity': grand_total_quantity,
        'grand_total_volume': grand_total_volume,
        'grand_total_area': grand_total_area,
        'url_name': 'lumber_track:report_to_stock_summary_result',
    }

    if request.GET.get('export') == 'excel':
        return export_summary_to_excel(summary_data, context['title'], date_from, date_to,
                                       grand_total_quantity, grand_total_volume, grand_total_area)

    return render(request, 'lumber_track/report_summary.html', context)


def report_to_shop_summary(request):
    """Сводный отчет по отгрузкам в магазин (форма выбора периода)"""
    if request.GET.get('date_from') and request.GET.get('date_to'):
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        return redirect(f'/reports/to-shop-summary/result/?date_from={date_from}&date_to={date_to}')

    return render(request, 'lumber_track/report_form.html', {
        'title': 'Магазин в Красноярске (сводный)',
        'url_name': 'lumber_track:report_to_shop_summary_result'
    })


def report_to_shop_summary_result(request):
    """Результат сводного отчета по отгрузкам в магазин"""
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if not date_from or not date_to:
        return redirect('lumber_track:report_to_shop_summary')

    from datetime import datetime
    from collections import defaultdict

    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()

    items = DocumentItem.objects.filter(
        document__doc_type=3,
        document__to_location_id=3,
        document__doc_date__gte=date_from_obj,
        document__doc_date__lte=date_to_obj
    ).select_related('product_name', 'species', 'grade', 'lumber_dim', 'unit_dim')

    report_data = defaultdict(lambda: defaultdict(list))

    for item in items:
        species = item.species.name
        grade = item.grade.code
        product_name = item.product_name.name

        if item.lumber_dim:
            dimension_display = f"{item.lumber_dim.thickness}×{item.lumber_dim.width}×{item.lumber_dim.length}"
        else:
            dimension_display = "—"

        items_list = report_data[species][grade]
        found = False
        for existing in items_list:
            if existing['product_name'] == product_name and existing['dimension_display'] == dimension_display:
                existing['quantity'] += item.quantity
                existing['volume'] += item.volume_m3
                existing['area'] += item.area_m2
                found = True
                break

        if not found:
            items_list.append({
                'product_name': product_name,
                'dimension_display': dimension_display,
                'quantity': item.quantity,
                'volume': item.volume_m3,
                'area': item.area_m2,
            })

    summary_data = {}
    grand_total_quantity = 0
    grand_total_volume = 0
    grand_total_area = 0

    for species, grades in report_data.items():
        summary_data[species] = {}
        for grade, items_list in grades.items():
            grade_total_quantity = sum(item['quantity'] for item in items_list)
            grade_total_volume = sum(item['volume'] for item in items_list)
            grade_total_area = sum(item['area'] for item in items_list)

            summary_data[species][grade] = {
                'items': items_list,
                'total_quantity': grade_total_quantity,
                'total_volume': grade_total_volume,
                'total_area': grade_total_area,
            }

            grand_total_quantity += grade_total_quantity
            grand_total_volume += grade_total_volume
            grand_total_area += grade_total_area

    context = {
        'title': 'Магазин в Красноярске (сводный)',
        'date_from': date_from_obj,
        'date_to': date_to_obj,
        'summary_data': summary_data,
        'grand_total_quantity': grand_total_quantity,
        'grand_total_volume': grand_total_volume,
        'grand_total_area': grand_total_area,
        'url_name': 'lumber_track:report_to_shop_summary_result',
    }

    if request.GET.get('export') == 'excel':
        return export_summary_to_excel(summary_data, context['title'], date_from, date_to,
                                       grand_total_quantity, grand_total_volume, grand_total_area)

    return render(request, 'lumber_track/report_summary.html', context)


def export_summary_to_excel(summary_data, title, date_from, date_to,
                            grand_total_quantity, grand_total_volume, grand_total_area):
    """Экспорт сводного отчета в Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    species_fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')
    grade_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовки
    headers = ['Наименование', 'Размер', 'Количество (шт)', 'Объем (м³)', 'Площадь (м²)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = thin_border

    current_row = 2

    for species, grades in summary_data.items():
        # Заголовок породы
        cell = ws.cell(row=current_row, column=1, value=f"🌲 {species}")
        cell.font = Font(bold=True, size=11)
        cell.fill = species_fill
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        for col in range(1, 6):
            ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1

        for grade, data in grades.items():
            # Заголовок категории
            cell = ws.cell(row=current_row, column=1, value=f"⭐ Категория {grade}")
            cell.font = Font(bold=True, size=10)
            cell.fill = grade_fill
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            for col in range(1, 6):
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1

            for item in data['items']:
                ws.cell(row=current_row, column=1, value=item['product_name'])
                ws.cell(row=current_row, column=2, value=item['dimension_display'])
                ws.cell(row=current_row, column=3, value=float(item['quantity']))
                ws.cell(row=current_row, column=4, value=float(item['volume']))
                ws.cell(row=current_row, column=5, value=float(item['area']))

                for col in range(1, 6):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = thin_border
                    if col >= 3:
                        cell.alignment = Alignment(horizontal='right')
                current_row += 1

            # Итог по категории
            cell = ws.cell(row=current_row, column=1, value=f"Итого по категории {grade}:")
            cell.font = Font(bold=True)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            ws.cell(row=current_row, column=3, value=float(data['total_quantity']))
            ws.cell(row=current_row, column=3).font = Font(bold=True)
            ws.cell(row=current_row, column=4, value=float(data['total_volume']))
            ws.cell(row=current_row, column=4).font = Font(bold=True)
            ws.cell(row=current_row, column=5, value=float(data['total_area']))
            ws.cell(row=current_row, column=5).font = Font(bold=True)
            for col in range(1, 6):
                ws.cell(row=current_row, column=col).border = thin_border
                if col >= 3:
                    ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='right')
            current_row += 1

    # Общие итоги
    ws.cell(row=current_row, column=1, value="ВСЕГО за период:")
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    ws.cell(row=current_row, column=3, value=float(grand_total_quantity))
    ws.cell(row=current_row, column=3).font = Font(bold=True)
    ws.cell(row=current_row, column=4, value=float(grand_total_volume))
    ws.cell(row=current_row, column=4).font = Font(bold=True)
    ws.cell(row=current_row, column=5, value=float(grand_total_area))
    ws.cell(row=current_row, column=5).font = Font(bold=True)
    for col in range(1, 6):
        ws.cell(row=current_row, column=col).border = thin_border
        if col >= 3:
            ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='right')

    # Ширина колонок
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{title}_{date_from}_{date_to}.xlsx"'
    wb.save(response)
    return response